#! Python 3
# - Copy and Paste Ranges using OpenPyXl library

import openpyxl as xl
from copy import copy
from openpyxl import workbook

import os
class Edite():
    """this class provide editing tools for excel"""
    def __init__(self,folder,readfile1,sheet1,row,column,writefile,writesheet):
        self.folder=folder
        
        self.readfile1=readfile1
        self.sheet1=sheet1
        self.row=row
        self.column=column
        
        self.writefile=writefile
        self.writesheet=writesheet

    def rename_sheets(self):
        '''to rename excel sheets '''
        os.chdir(self.folder)
        
        wb1 = xl.load_workbook(self.readfile1)
        ws1 = wb1.get_sheet_by_name(self.sheet1) 
        ws1.title=(self.writesheet)
        wb1.save(self.readfile1)
    
    def deletrows(self):
        '''to delete rows from sheets '''
        os.chdir(self.folder)
        
        wb1 = xl.load_workbook(self.readfile1)
        ws1 = wb1.get_sheet_by_name(self.sheet1) 
        ws1.delete_rows(1)
        wb1.save(self.readfile1)        
    def delete_word():
    
        for dirpath, dirnames, filenames in os.walk(path):
            for f in filenames:  
                if f.endswith(".docx") :
                
                    os.remove(os.path.join(dirpath,f))

                if f.endswith(".doc"):
                
                    os.remove(os.path.join(dirpath,f))
                if f.endswith(".xlsx"):
                    os.remove(os.path.join(dirpath,f))

                if f.endswith(".xls"):
                    os.remove(os.path.join(dirpath,f))


    def copy_cells(self):
        os.chdir(self.folder)
        path=self.folder
        default_sheet=xl.load_workbook("standard_format.xlsx")
        for dirpath, dirnames, filenames in os.walk(path):
            print(filenames)
            
            for f in filenames:  
                if f.endswith(".xlsx") :
                    in_file =(dirpath + '/'+ f)
                    wb1 = xl.load_workbook(in_file)          
                    
                    for row in default_sheet.rows :
                        for cell in row:
                            new_cell = wb1.cell(row=cell.row, column=cell.col_idx,
                                    value= cell.value)
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
                if f.endswith(".xls") :
                    in_file =(dirpath + '/'+ f)
                    wb1 = xl.load_workbook(in_file)
                    
                    default_sheet = workbook[f]
                    for row in default_sheet.rows:
                        for cell in row:
                            new_cell = wb1.cell(row=cell.row, column=cell.col_idx,
                                    value= cell.value)
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
