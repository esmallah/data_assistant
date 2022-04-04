'''this module for the data was emported from excel sheet analysis'''

import pandas as pd
import openpyxl as xl
from openpyxl import load_workbook

import numpy as np
import os
from copy import copy
import random
from random import randint,seed
from openpyxl.chart import BarChart, Reference, Series,LineChart
from .collect import columns_quality as qc_molds
#cols=["day","machine","product_name","product_code","product_parts","standard_dry_weight",
#"standard_dry_weight_from","standard_dry_weight_to","average_dry_weight"
#,"standard_rate_hour","c_t_standard_per_second","set","rat_actually","c_t_actually","sum_scrabe_no_parts"
#,"number_scrab_by_item","gross_production_by_set","scrabe_standard"
#,"gross_production","scrap_percent_by_item","customer_id","mold_id","item_id"]
        

col_rename={"product_name": "اسم المنتج","product_name_by_parts":"اسم المنتج بالاجزاء"
        ,"product_code":"كود المنتج","product_parts":"اجزاء المنتج","machine_id":'رقم الماكينة',"number_day_use":"عدد ايام التشغيل"
        ,"standard_dry_weight_from":"مواصفة الوزن الجاف من","standard_dry_weight_to":"مواصفة الوزن الجاف إلي",
        "weight_under_validation":"الاوزان الاقل من المواصفة","weight_above_validation":"الاوزان الاعلي من المواصفة",
        "c_t_deviation":"الفارق بين المعدل المعيارى والفعلى","rat_standard_deviation":'الانحراف المعياري لمعدل الانتاج',
        'rat_validation':'التحقق من مطابقة معدل الانتاج للمواصفة',
        "sum_scrabe_no_parts":"عدد التوالف بالقطع","number_scrab_by_item":"عدد التوالف بالطقم",
        "gross_production_by_set":"إجمالي الانتاج بالطقم","gross_production":"إجمالي الانتاج بالقطعة","scrap_percent_by_item":"نسبة الاسكراب بالطقم",
        "product_groub":"اسم المنتج المجمع","standard_dry_weight":"الوزن الجاف المعياري","average_dry_weight":"متوسط الوزن الجاف الفعلي"
        ,"standard_rate_hour":"المعدل المعياري بالساعة","rat_actually":"متوسط المعدل الفعلي بالساعة"
        ,"c_t_standard_per_second":"زمن الدورة المعياري بالثانية","c_t_actually":"متوسط زمن الدورة الفعلي بالثانية"
        ,"scrap_standard":"معياري الاسكراب","parts_patchsNumbers":"ارقام الباتشات بالجزء","scrap_percent_by_item":"نسبة التوالف بالقطعة"
        ,"Items_patchsNumbers":"ارقام الباتشات بالاصناف","scrabe_standard":"معياري التوالف","part_id":"رقم الجزء",
        "mold_id":"رقم الاسطمبة","item_id":"رقم المنتج" }      # for recaull module to change english name to arabic

columns_weight=["day","year","month","machine_id","product_name_by_parts","product_name","product_code","standard_dry_weight",
"standard_dry_weight_from","standard_dry_weight_to",'shift1_dry_weight1','shift1_dry_weight2','shift1_dry_weight3',
        'shift1_dry_weight4','shift1_dry_weight5','shift2_dry_weight1','shift2_dry_weight2',
        'shift2_dry_weight3','shift2_dry_weight4','shift2_dry_weight5',"average_dry_weight","part_id",'factory','deepth_mm','id_DayPartUnique']

columns_machine=["day","year","month","machine_id","product_name","product_code","scrabe_standard","sum_scrabe_no_parts"
,"number_scrab_by_item","gross_production",'gross_production_by_set','parts_patchsNumbers','scrap_percent_by_item'
,"number_day_use",'Items_patchsNumbers',"item_id","machine_type",'factory','tall_mm','id_DayPartUnique']

columns_cycle_time=["day","year","month","machine_id","mold_name","product_name","product_code","set","standard_rate_hour"
,"c_t_standard_per_second",'shift1_c_t1','shift1_c_t2','shift2_c_t1','shift2_c_t2',"rat_actually","c_t_actually","number_day_use",'rat_validation',"mold_id",'factory','width_mm','id_DayPartUnique']
        
class Group():
    def __init__(self,folder,readfile,readsheet,column1,column2,writefile,writesheet):
        #self.folder=folder
        self.folder=folder
        self.readfile=readfile
        self.readsheet=readsheet
        self.column1=column1
        self.column2=column2
        self.writefile=writefile
        self.writesheet=writesheet

    def monthly(self):
        os.chdir(self.folder)

        reader_file=self.readfile
    
        daily_analysis3= pd.read_excel(reader_file,self.readsheet)
        daily_analysis_bool3=daily_analysis3["year"]==self.column1
        daily_analysis2=daily_analysis3[daily_analysis_bool3]
        daily_analysis_bool2=daily_analysis2["month"]==self.column2
        daily_analysis=daily_analysis2[daily_analysis_bool2]
        
        daily_analysis_bool=daily_analysis["day"]<=self.writefile #for less than the day
        daily_analysis=daily_analysis[daily_analysis_bool]
        daily_analysis=daily_analysis[qc_molds]       
       
        dry_weight3=daily_analysis[columns_weight]
        dry_weight_bool=dry_weight3['average_dry_weight'].notnull()
        dry_weight2=dry_weight3[dry_weight_bool]

        scrap4=daily_analysis[columns_machine]
        scrap_bool=scrap4["gross_production"].notnull()
        scrap3=scrap4[scrap_bool]
        scrap2=scrap3
        molds_rate3=daily_analysis[columns_cycle_time]
        molds_rate_bool=molds_rate3["c_t_actually"].notnull()
        molds_rate2=molds_rate3[molds_rate_bool]
        
        #scrap
        scrap=scrap2.groupby(["product_name","product_code","scrabe_standard"])['number_scrab_by_item',
        "gross_production_by_set","sum_scrabe_no_parts","gross_production","number_day_use"].sum()
        
        scrap["scrap_percent_by_item"]=scrap['number_scrab_by_item']/scrap['gross_production']
        #scrap= pd.DataFrame()
        

        scrap_product_machine=scrap2.groupby(["product_name","product_code","machine_id","scrabe_standard"])['number_scrab_by_item',
        "gross_production_by_set","sum_scrabe_no_parts","gross_production","number_day_use"].sum()
        scrap_product_machine["scrap_percent_by_item"]=scrap_product_machine['number_scrab_by_item']/scrap_product_machine['gross_production']
        

        scrap_machine_product=scrap2.groupby(["machine_id","scrabe_standard","product_name","product_code"])['number_scrab_by_item',
        "gross_production_by_set","sum_scrabe_no_parts","gross_production","number_day_use"].sum()
        scrap_machine_product["scrap_percent_by_item"]=scrap_machine_product['number_scrab_by_item']/scrap_machine_product['gross_production']
        

        machines=scrap2.groupby(["machine_id","scrabe_standard","machine_type"])['number_scrab_by_item',"gross_production_by_set",
        'sum_scrabe_no_parts',"gross_production","number_day_use"].sum()
        machines["scrap_percent_by_item"]=machines['number_scrab_by_item']/machines['gross_production']
        scrap_bool2=scrap2["scrap_percent_by_item"]>scrap2["scrabe_standard"]
        scrap_ncr=scrap2[scrap_bool2]
        print("please weight")
        #production rate report
        
        molds_rate=molds_rate2.groupby(["product_name","product_code","set","standard_rate_hour"
        ,"c_t_standard_per_second"])["rat_actually","c_t_actually"].mean()
        #molds_rate_bools_ncr=molds_rate["rat_actually"]>molds_rate["standard_rate_hour"]
        #molds_rate_ncr=molds_rate[molds_rate_bools_ncr]
        dry_weight=dry_weight2.groupby(["product_name_by_parts","product_code","standard_dry_weight",
        "standard_dry_weight_from","standard_dry_weight_to"])["average_dry_weight"].mean()
        #dry_weight_bool_low=dry_weight["average_dry_weight"] < dry_weight["standard_dry_weight_from"]
        #dry_weight_bool_high=dry_weight["average_dry_weight"] > dry_weight["standard_dry_weight_to"]
        #dry_weight_ncr_low=dry_weight[dry_weight_bool_low]
        #dry_weight_ncr_high=dry_weigh[dry_weight_bool_high]
        # export
        

        #validation data
        #scap validation
        product_parts_input=daily_analysis["gross_production"].sum
        
        
        scrab_set_input=daily_analysis["number_scrab_by_item"].sum
        #scrab_input=product_parts_input + product_set_input + scrab_parts_input + scrab_set_input

        product_parts_outbut=machines["gross_production"].sum
        
        
        scrab_set_outbut=machines["number_scrab_by_item"].sum
        #scrab_output=product_parts_outbut + product_set_outbut +  scrab_parts_outbut + scrab_set_outbut

        #if  scrab_input!=scrab_output:#not equal validation
        #    exit
        #    print("the scrap data or production data is worng kinldy review ")
        #export analysis to excel sheets is named output in the same folder
        
                # export monthly
                #  analysis 
        writer = pd.ExcelWriter("QC_molds_monthly_until.xlsx")
        daily_analysis.rename(columns={c:c.lower() for c in col_rename})      

        daily_analysis.to_excel(writer,'input', index=False)

        scrap.rename(columns={c:c.lower() for c in col_rename})
        scrap.to_excel(writer,'scrap_product', index=True)

        scrap_product_machine.rename(columns={c:c.lower() for c in col_rename})
        scrap_product_machine.to_excel(writer,'scrap_product_machines', index=True)

        scrap_machine_product.rename(columns={c:c.lower() for c in col_rename})
        scrap_machine_product.to_excel(writer,'scrap_machines_product', index=True)
        
        machines.rename(columns={c:c.lower() for c in col_rename})
        machines.to_excel(writer,"scrap_machines", index=True)
        
        dry_weight.rename(columns={c:c.lower() for c in col_rename})
        dry_weight.to_excel(writer,'weights', index=True)
        molds_rate.rename(columns={c:c.lower() for c in col_rename})
        molds_rate.to_excel(writer,'c_t', index=True)
        #weight_ncr_low=weight_ncr_low.rename(index=str, columns=col_rename)
        #weight_ncr_low.to_excel(writer,'weight_ncr_low')
        #weight_ncr_high=weight_ncr_high.rename(index=str, columns=col_rename)

        
        
        print ("analysis for ")
        print(self.column1)
        print(self.column2)
        print("for the days")
        print(daily_analysis["day"].unique())
        writer.save()
                    
    def daily_molds(self,yearDb,monthDb,dayDb):
        from memory import Block,cursor,conn

        os.chdir(self.folder)
        
        Block.get_daily_dataentry_items(self,yearDb,monthDb,dayDb)    
        get_data2=cursor.fetchall()
        #__________________________________________________________________        
        column_names = [desc[0] for desc in cursor.description]
        mold_analysis4 = pd.DataFrame(get_data2,columns=column_names)

        
        last_year=int(mold_analysis4["year"].max())
        print("_________daily report___________for yearr________",last_year,type(last_year))       
        mold_analysis_bool4=mold_analysis4["year"]==last_year
        mold_analysis3=mold_analysis4[mold_analysis_bool4]
        
        last_month=int(mold_analysis3["month"].max())
        mold_analysis_bool3=mold_analysis3["month"]==last_month
        mold_analysis2=mold_analysis3[mold_analysis_bool3]
        
        #dateDay3=mold_analysis2['date_day'].tail(1)
        print("_________daily report___________for month________",last_month,type(last_month))       
        #convert selecting to value
        #dateDay2=dateDay3.values
        #convert numpy ndarray to list
        #dateDay1 = dateDay2.tolist()
        #convert list  to string
        #dateDay = ' '.join([str(elem) for elem in dateDay1]) 
        #day=int(dateDay.split("/",2)[1])

        mold_days=mold_analysis2["day"].count()
        day=int(dayDb)
        daily_analysis1_bool=mold_analysis2["day"]==day
        daily_analysis1=mold_analysis2[daily_analysis1_bool]
        #validate input 
        


        #for fix nan error in ct
        daily_analysis = daily_analysis1.dropna(subset=['c_t_actually'])#remove all numric data
        daily_analysis = daily_analysis.dropna(subset=['c_t_actually'])#drop And for remove all rows with NaNs in column x use dropna: 
        daily_analysis["c_t_actually"]=daily_analysis["c_t_actually"].astype(int)#Last convert values to ints:

        report_forCt=daily_analysis.groupby(["machine_id","mold_id"])["standard_dry_weight_from","standard_dry_weight_to","c_t_standard_per_second","standard_rate_hour","average_dry_weight","rat_actually","c_t_actually"].mean()
        report=daily_analysis1.groupby(["machine_id","mold_id"])["standard_dry_weight_from","standard_dry_weight_to","average_dry_weight"].mean()
        

        print("___________report   data",report)
        #report=pd.DataFrame()
        
        
        mold_count=report_forCt["rat_actually"].count()
        #cycle timpe report
        
        c_t=report_forCt[["c_t_standard_per_second","standard_rate_hour","rat_actually","rat_validation","c_t_actually"]]
        
        c_t_nonconfomity=c_t[c_t['rat_actually']*1.05<c_t['standard_rate_hour']]
        c_t_nonconfomity_count=c_t_nonconfomity["rat_actually"].count()
        
        
        if c_t_nonconfomity_count==0:  # for ignor impty index error
            c_t_nonconfomity=pd.DataFrame(index=[0])
            c_t_nonconfomity["mold_id"]=0
            c_t_nonconfomity["c_t_standard_per_second"]=0
            c_t_nonconfomity["standard_rate_hour"]=0
            c_t_nonconfomity["rat_actually"]=0
            c_t_nonconfomity["rat_validation"]=0
            c_t_nonconfomity["rat_validation"]=0
            
        ct_ok=mold_count-c_t_nonconfomity_count
        c_t_nonconfomity["c_t_nonconfomity_count"]=c_t_nonconfomity_count
        c_t_nonconfomity["ct_ok"]=ct_ok
        #weight report
       
        
        wieght2=report
        wieght=wieght2[wieght2["dryweight_deviation_validation"]<1]#filter nonconformity data
        print("___________weight   data",wieght)
        #filter low weithrs
        weight_nonconfomity_low=wieght[wieght['average_dry_weight']<wieght["standard_dry_weight_from"]] #add column tocount number of non conformity product
        
        #add rows for filter high weight
            #fix error for zero ncr
        #if daily_analysis1["dryweight_deviation_validation"]==0 
        weight_nonconfomity_high=wieght[wieght['average_dry_weight']>wieght["standard_dry_weight_to"]]
        #weight_nonconfomity=pd.DataFrame(index=[["machine_id","mold_name"],0])
        #weight_nonconfomity=pd.DataFrame(index=[wieght["machine_id","mold_name"]])
        weight_nonconfomity=weight_nonconfomity_high
        weight_nonconfomity_count=weight_nonconfomity["average_dry_weight"].count()
        if weight_nonconfomity_count==0:  # for ignor impty index error        
            weight_nonconfomity=pd.DataFrame(index=[0])
        if weight_nonconfomity_count>1:  # for ignor impty index error        
            weight_nonconfomity=pd.DataFrame()##error we muast select index of groupby
        
        weight_nonconfomity=weight_nonconfomity.append(weight_nonconfomity_low)    #for append the light weights
        weight_nonconfomity=weight_nonconfomity.append(weight_nonconfomity_high)    #for append the heavy weights
        
        weight_nonconfomity['weight_nonconfomity_count']=weight_nonconfomity_count
        weight_ok=mold_count-weight_nonconfomity_count
        weight_nonconfomity['weight_ok']=weight_ok
        
        ##screap report by parts##_________
                #scap by parts
        #report=pd.DataFrame()
        scrap5=daily_analysis1[["machine_type","machine_id",'number_scrab_by_item','gross_production','scrap_percent_by_item',"number_day_use","mold_id","scrap_weight_kg","production_weight_kg","product_name","scrabe_standard"]]
        
        #scrap5=report.append(scrap5)#we need fix rong calucate sacrap percent for 
        

        scrap_newmachines=scrap5[scrap5["machine_type"]=="new_machine"]
        scrap_oldmachines=scrap5[scrap5["machine_type"]=="old_machine"]
        

        scrap_part_new=scrap_newmachines["number_scrab_by_item"].sum()
        production_part_new=scrap_newmachines["gross_production"].sum()
        
        scrap_percent_new=scrap_part_new/production_part_new * 100
        scrap_part=scrap5["number_scrab_by_item"].sum()
        production_set=scrap5["gross_production"].sum()
        scrap_percent=scrap_part/production_set * 100
        scrap4=scrap5[scrap5["number_day_use"]==1]      #for filter the using machines only
        
    
        scrap3=scrap4[scrap4["scrap_percent_by_item"]>scrap4["scrabe_standard"]]
        
        scrap2=scrap3[scrap3["gross_production"]>scrap3["number_scrab_by_item"]]# for filter any machines didn't create production
        
        scrap=scrap2
        print("_______________scrap____________________",scrap)
        scrap["scrab_parts_new_machine"]=scrap_part_new
        
        scrap["production_parts_new_machine"]=production_part_new
        scrap["production_parts_all"]=production_set
        scrap["scrap_percent_new_machine"]=scrap_percent_new
        scrap["scrap_percent_all"]=scrap_percent
        scrap["scrab_parts_all"]=scrap_part
        scrap_nonconfomity_count=scrap["number_scrab_by_item"].count()
        
        ##scrap for molds___________________
        '''must be in excel sheet (input) not showing null value for not mistacks in average result'''
        scrap_molds=scrap5.groupby(["machine_type","machine_id","mold_id","scrabe_standard"])['number_scrab_by_item'].sum()
        scrap_molds["gross_production"]=scrap5.groupby(["machine_type","machine_id","mold_id","scrabe_standard"])['gross_production'].sum()
        scrap_molds["scrap_weight_kg"]=scrap5.groupby(["machine_type","machine_id","mold_id","scrabe_standard"])['scrap_weight_kg'].sum()
        scrap_molds["production_weight_kg"]=scrap5.groupby(["machine_type","machine_id","mold_id","scrabe_standard"])['production_weight_kg'].sum()
        scrap_molds["number_scrab_by_item"]=scrap5.groupby(["machine_type","machine_id","mold_id","scrabe_standard"])['number_scrab_by_item'].sum()
        production_set_molds=scrap_molds["gross_production"].sum()
        scrap_set_molds=scrap_molds["number_scrab_by_item"].sum()
        scrap_percent_molds=scrap_set_molds/production_set_molds*100
        scrap_molds["percent"]=(scrap_molds["number_scrab_by_item"]/scrap_molds["gross_production"])*100
        ##scrap for items(as item master)___
        scrap_items=scrap5.groupby(["machine_type","machine_id","product_name"])['number_scrab_by_item'].sum()
        
        scrap_items["gross_production"]=scrap5.groupby(["machine_type","machine_id","product_name"])['gross_production'].sum()
        scrap_items["number_day_use"]=scrap5.groupby(["machine_type","machine_id","product_name"])['number_day_use'].mean()
        scrap_items["scrap_weight_kg"]=scrap5.groupby(["machine_type","machine_id","product_name"])['scrap_weight_kg'].sum()
        scrap_items["production_weight_kg"]=scrap5.groupby(["machine_type","machine_id","product_name"])['production_weight_kg'].sum()
        scrap_items["number_scrab_by_item"]=scrap5.groupby(["machine_type","machine_id","product_name"])['number_scrab_by_item'].sum()
        scrap_items["percent"]=((scrap_items["number_scrab_by_item"])/(scrap_items["gross_production"]))*100
#        molds_newmachines=scrap_molds.filter(lambda x: x['machine_type']=="new_machine")
        
        
        
        scrap_percent_new=scrap_part_new/production_part_new * 100
        ##scrap non conformity_______________________
        
        
        if scrap_nonconfomity_count>0:  # for ignor impty index error
        #    scrap=pd.DataFrame(index=[0])
            scrap_nonconfomity=scrap2.groupby(["machine_id","mold_id"])['number_scrab_by_item'].sum()
            #for ignor error when srcap is 0
            
            
        else:
            #if scrap_nonconfomity.datatypt==0:  # for ignor impty index error
            scrap_nonconfomity=pd.DataFrame(index=[0])
            scrap_nonconfomity["mold_name"]=0
        print("_____________test__________scrap nonconfromity____",scrap_nonconfomity)
        #extract excel report
        writer_report = pd.ExcelWriter("QC_molds_daily_archive.xlsx")
        mold_analysis2.to_excel(writer_report,"input", index=False)
        writer_report.save()

        writer = pd.ExcelWriter("day_analysis2.xlsx")

        #weight_nonconfomity.to_excel(writer,"weight_ncr")
        
        c_t_nonconfomity.to_excel(writer,"c.t")
        scrap.to_excel(writer,"scrap")
        scrap_molds.to_excel(writer,"scrap_molds",merge_cells=False)
        scrap_nonconfomity.to_excel(writer,"scrap_ncr")
        scrap_items.to_excel(writer,"scrap_items",merge_cells=False)
    
        report.to_excel(writer,merge_cells=False)
        daily_analysis1.to_excel(writer,"input_molds", index=False)
        
        writer.save()
        
        #extract daily excel report by formating

        wb = load_workbook("QC_molds_daily_summary.xlsx")
        ws2= wb.get_sheet_by_name('daily')
        
        ws=wb.copy_worksheet(ws2)   #copy new sheet

        #splite day
        #lastDay = [daily_analysis1["day_date"].split('-')[2] for d in daily_analysis1.Date]
        #lastDay=pd.to_DateFrame(daily_analysis1,columns=["day_date"])
        #daily_analysis1.lastDay=pd.to_datetime(lastDay.Datetime,format='%d')
        
        ws.title=str(day)  #rename new sheet by the name
        
                
        # Data can be assigned directly to cells
        ws['b1'] = day            #day for molds report
        ws['c1'] = last_month    #month for molds report
        ws['d1'] = last_year    #year for molds report
        ws['A4'] = scrap_percent           #scrap on all machie
        ws['B4'] = scrap_percent_new    #scrap on new macine
        
        ws['a14'] =c_t_nonconfomity.iloc[0][6]                #ct pass 
        ws['b14'] =c_t_nonconfomity.iloc[0][5]                #ct not acceptable
        
        ws['a25'] = weight_nonconfomity.iloc[0][5]             #weight pass 
        ws['b25'] =weight_nonconfomity.iloc[0][4]
        #if weight_nonconfomity_high>=1:  # for ignor impty index error        
        #    ws['b25'] =weight_nonconfomity_low.iloc[0][4]            #weights low not acceptable
        #else:
        #    ws['b25'] =0
        #ws['a35'] = weight_nonconfomity.iloc[0][5]             #weight pass 
        #if weight_nonconfomity_high>=1:  # for ignor impty index error        
        #    ws['b35'] =weight_nonconfomity_high.iloc[0][4]            #weights hig not acceptable
        #else:
        #    ws['b35'] =0           #weights hig not acceptable
        #_______
        
        ws['a35'] = weight_nonconfomity.iloc[0][5]             #weight pass 
        ws['b25'] =weight_nonconfomity.iloc[0][4]

        #to select index of columns for scraps
        if scrap_nonconfomity_count>=1:
            rows = scrap_nonconfomity.index
            r = 4  # start at 10th row
            c = 3 # column 'c'
            for row in rows:       
                for item in row:
                    ws.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 3
                r += 1

        #to select index of columns to weights
        
        if c_t_nonconfomity_count>=1:  # for ignor impty index error        
            rows = c_t_nonconfomity.index
            #rows = weight_nonconfomity.index
            r = 14  # start at 10th row
            c = 3 # column 'c'
            
            for row in rows:       
                for item in row:
                    ws.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 3
                r += 1
        
        #to select index of columns to light weights 
        if weight_nonconfomity_count>=1:  # for ignor impty index error        
            rows = weight_nonconfomity_low.index
            #rows = weight_nonconfomity.index
            r = 25  # start at 10th row
            c = 3 # column 'c'
            
            for row in rows:       
                for item in row:
                    ws.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 3
                r += 1
        #to select index of columns to hight weights 
        if weight_nonconfomity_count>=1:  # for ignor impty index error        
            rows = weight_nonconfomity_high.index
            #rows = weight_nonconfomity.index
            r = 35  # start at 10th row
            c = 3 # column 'c'
            
            for row in rows:       
                for item in row:
                    ws.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 3
                r += 1
        
        wb.save("QC_molds_daily_summary.xlsx")
        #__________________archinve work boook monthly and daily report________________________________________        
        #__________________archinve work boook monthly and daily report general________________________________________
        #input report
        wb_formats=load_workbook("format_QC_reports_v2.xlsx")    
        ws_input= wb_formats.get_sheet_by_name('input')
        #ws_input=wb_formats.copy_worksheet(ws_input2)   #copy new sheet
        ws_input.sheet_view.zoomScale = 60 #zoom set

        column_size_input=pd.DataFrame(mold_analysis2,columns=["month"]).shape[0]
        #print("______test___")
        #print(mold_analysis2)
        
        r = 4  # start at 4th row
        c = 1 # column 'a'
        
        for row in range(0,column_size_input):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
            rows = mold_analysis2.iloc[row]
            for item in rows:
                ws_input.cell(row=r, column=c).value = item
                c += 1 # Column 'd'
            c = 1
            r += 1
        #__________________archinve work boook monthly and daily report for specific mold
        
        
        wb_formats.save("QC_molds_daily_archive.xlsx")    
    
    def copy_between_workbooks(self):        
        print("________starting copy last sheet____")
        import win32com
        from win32com.client import DispatchEx
        os.chdir(self.folder)
        print("______________test_________")
        print(self.folder)
        #input last day sheets
        
        excel = DispatchEx('Excel.Application')
        wbP=excel.Workbooks.Open('QC_molds_daily_archive')
        wbG=excel.Workbooks.Open(self.readfile)
        #wbP=excel.Workbooks.Open(r'C:\Temp\Junk\Temp.xlsx')
        #wbG=excel.Workbooks.Open(r'C:\Temp\Junk\Temp2.xlsx')
        wbG.Worksheets("deleted").Copy(Before=wbP.Worksheets(self.readsheet))
        wbP.SaveAs(r'C:\Temp\Junk\Temp.xlsx')
        excel.Quit()
        del excel # ensure Excel process ends
    def copy_between_workbooks2(self,worksheet,index=None):        
        print("get last sheet in day")
        os.chdir(self.folder)

        wb=load_workbook(self.readfile)
     #   def add_sheet(self, worksheet, index=None):
        """Add an existing worksheet (at an optional index)."""
#        if not isinstance(worksheet, self._worksheet_class):
 #           raise TypeError("The parameter you have given is not of the type '%s'" % self._worksheet_class.__name__)

        #if index is None:
        #   self.worksheets.append(worksheet)
        #else:
        #self.worksheets.insert(index, worksheet)
        ws_input= wb.get_sheet_by_name(worksheet)
        wb_formats=load_workbook("QC_molds_daily_archive.xlsx")    
        #if not isinstance(worksheet, self._worksheet_class):
        #    raise TypeError("The parameter you have given is not of the type '%s'" % self._worksheet_class.__name__)

        if index is None:
            wb_formats.worksheets.append(ws_input)
        else:
            wb_formats.worksheets.insert(index, worksheet)
        wb_formats.worksheets.insert(1, ws_input)

        wb_formats.save("QC_molds_daily_archive.xlsx")    

    
    def yearly(self):#for statiscs to qc yearly molds report
        os.chdir(self.folder)
        reader=pd.read_excel(self.readfile,self.readsheet)
        #for more filtraation about id_molds
        
        input_report3=reader[qc_molds]
        #input_report2=input_report3[input_report3["customer_name"]==customer]       
        input_report2=input_report3

        input_report=input_report2[input_report2['year']==self.column1]
        #verification 
        #1 for the right ct? ( previous mistaks is bad input)
        #calcuate the logic actioly rat by recalculat form actioly cycltype
        #????????
        #2 for the right names? ( previous mistaks is bad input)
        #calcuate the simulation between names
        #????????
        #ct report
        clos_cycle2=input_report[columns_cycle_time]
        ct_bool=clos_cycle2["c_t_actually"].notnull()
        clos_cycle=clos_cycle2[ct_bool]
        df = pd.DataFrame(clos_cycle, columns = qc_molds )
        rat=pd.crosstab([df.mold_name,df.standard_rate_hour],[df.year,df.month],df.rat_actually,aggfunc='mean',margins=False)
        ct=pd.crosstab([df.mold_name,df.c_t_standard_per_second],[df.year,df.month],df.c_t_actually,aggfunc='mean',margins=False)
        
        #weight report
        clos_weight2=input_report[columns_weight]
        weight_bool=clos_weight2["average_dry_weight"].notnull()
        clos_weight=clos_weight2[weight_bool]
        df2 = pd.DataFrame(clos_weight, columns = qc_molds )

        weight=pd.crosstab([df2.product_name_by_parts,df2.standard_dry_weight,
        df2.standard_dry_weight_from,df2.standard_dry_weight_to],[df2.year,df2.month],df2.average_dry_weight
        ,aggfunc='mean',margins=False)
        
        print("pleas wait")
        #scrap report
        scrap2=input_report[columns_machine]
        scrap_bool=scrap2['gross_production'].notnull()
        scrap=scrap2[scrap_bool]
        
        df3 = pd.DataFrame(scrap, columns = qc_molds )
        
        scrap_set=pd.crosstab([df3.product_name,df3.product_code,df3.scrabe_standard],[df3.year,df3.month],df3.number_scrab_by_item
        ,aggfunc='sum',margins=False)
        
        scrap_parts=pd.crosstab([df3.product_name,df3.product_code,df3.scrabe_standard],[df3.year,df3.month],df3.sum_scrabe_no_parts
        ,aggfunc='sum',margins=False)
        production_set=pd.crosstab([df3.product_name,df3.product_code,df3.scrabe_standard],[df3.year,df3.month],df3.gross_production
        ,aggfunc='sum',margins=False)
        production_parts=pd.crosstab([df3.product_name,df3.product_code,df3.scrabe_standard],[df3.year,df3.month],df3.gross_production
        ,aggfunc='sum',margins=False)
        
        

        #final report
        #day_use=reader["number_day_use"]
        #pd.to_numeric(day_use, errors='ignore')
        #report=clos_cycle.groupby(["mold_name"])["number_day_use"].sum()

        #report["product_name"]=weight["product_name"]
        #report["number_day_use"]=day_use
        
        #ouut put by rename columns
        writer = pd.ExcelWriter(self.writefile)
        input_report=input_report.rename(index=str, columns=col_rename)
        input_report.to_excel(writer,"input",index=False)

        rat=rat.rename(index=str, columns=col_rename)
        rat.to_excel(writer,"rat")
        
        ct=ct.rename(index=str, columns=col_rename)
        ct.to_excel(writer,"ct")
        
        weight=weight.rename(index=str, columns=col_rename)
        weight.to_excel(writer,"weight")
        print("gust wait")
        
        scrap_set=scrap_set.rename(index=str, columns=col_rename)
        scrap_set.to_excel(writer,'scrap_set')
        
        scrap_parts=scrap_parts.rename(index=str, columns=col_rename)
        scrap_parts.to_excel(writer,'scrap_parts')

        production_set=production_set.rename(index=str, columns=col_rename)
        production_set.to_excel(writer,'production_set')
        production_parts=production_parts.rename(index=str, columns=col_rename)
        production_parts.to_excel(writer,'production_parts')
        #report=report.rename(index=str, columns=col_rename)
        #report.to_excel(writer,'report')
        
        clos_weight.to_excel(writer,'input_weight')
        clos_cycle.to_excel(writer,'input_ct')
        scrap.to_excel(writer,'input_scrap')
        writer.save()
    
    def yearly_ct(self):#for statiscs to qc molds report
        os.chdir(self.folder)
        reader=pd.read_excel(self.readfile,self.readsheet)
        #input_report=reader[qc_molds]
        #verification 
        #1 for the right ct? ( previous mistaks is bad input)
        #calcuate the logic actioly rat by recalculat form actioly cycltype
        #????????
        #2 for the right names? ( previous mistaks is bad input)
        #calcuate the simulation between names
        #????????
        #ct report
        clos_cycle2=reader[columns_cycle_time]
        ct_bool=clos_cycle2["c_t_actually"].notnull()
        clos_cycle=clos_cycle2[ct_bool]
        df = pd.DataFrame(clos_cycle, columns = columns_cycle_time )
        rat=pd.crosstab([df.mold_name,df.standard_rate_hour],[df.year,df.month],df.rat_actually,aggfunc='mean',margins=False)
        ct=pd.crosstab([df.mold_name,df.c_t_standard_per_second],[df.year,df.month],df.c_t_actually,aggfunc='mean',margins=False)
        

        #final report
        
        #ouut put by rename columns
        writer = pd.ExcelWriter(self.writefile)
        
        rat=rat.rename(index=str, columns=col_rename)
        rat.to_excel(writer,"rat")
        
        ct=ct.rename(index=str, columns=col_rename)
        ct.to_excel(writer,"ct")
        
        
        
        clos_cycle.to_excel(writer,'input_ct')
        
        writer.save()
    def generate_false_data(self,column1_number,column2_number,row_numbers,cell_form,cell_to,column_numbers):
        os.chdir(self.folder)

        reader=pd.read_excel(self.readfile,self.readsheet)
        
        wb = xl.load_workbook(self.readfile)
        ws= wb.get_sheet_by_name(self.readsheet)
        
        #column_size=reader["month"].shape
        column_size=pd.DataFrame(reader,columns=["month"]).shape[0]
        int(column_size)

        #reader=reader2.iloc[1:column_size+3]#select from 2nd row to last of rows
        #ws2=list(ws.rows)[2]
        
        for i in range(2,column_size):
                       
            LSL=reader.iloc[i][column1_number]#lower limit (row i and column G)
            int(round(LSL,0))
            
            USL=reader.iloc[i][column2_number]#upper limit (row i and column H)
            int(round(USL,0))
            
            weight_average=(LSL+ USL)/2 #average
            differance=(weight_average-LSL)/3 #for decrease diviation
            LCL=weight_average-differance  #for rais CPK
            
            UCL=weight_average+differance  #for rais CPK
              
            
            c = 9 # column 'H'
            for n in range(row_numbers):
                #random.seed(10)
                random_value = random.uniform(LCL,UCL)  
                #random_value = randint(LSL,USL)  
                #str(round(random_value,0))          
                #int(random_value)
                #values.append(random_value)

    #                for item in values:
                
                ws.cell(row=i, column=c).value = random_value
                c += 1
                #n+=1
            
        wb.save(self.writefile)

    def spc_molds(self,year,month,*mold,create_workbook=True):
            '''this functions for select vba anlaysis '''
            print("________starting statsic process charts____")
            #genecal data
            os.chdir(self.folder)
            
            #prepare sheets
            
            
            
            
            #ws_index=wb_formats.copy_worksheet(ws_index)
            save_name=str(year)+"-"+str(month)+" QC_SPC.xlsx"
            #prepare data
            if create_workbook:
                wb_formats=load_workbook("format_QC_reports_v2.xlsx")
                ws_data2=wb_formats.get_sheet_by_name("spc")

                ws_data=wb_formats.copy_worksheet(ws_data2)
                ws_data.title=str(mold)  

            else:
                wb_formats=load_workbook(save_name)
                ws_data2=wb_formats.get_sheet_by_name("spc")

                ws_data=wb_formats.copy_worksheet(ws_data2)
                ws_data.title=str(mold)  
            
            ws_index=wb_formats.get_sheet_by_name("index")
            cols_monthly=["machine_id","mold_name","product_name","product_code","standard_dry_weight",
            "standard_dry_weight_from","standard_dry_weight_to","average_dry_weight","year","month","day",
            "standard_rate_hour","c_t_standard_per_second","c_t_actually","rat_actually","item_id","mold_id","scrabe_standard",'sum_scrabe_shortage_bySet','sum_scrabe_roll_bySet',
            'sum_scrabe_broken_bySet','sum_scrabe_curve_bySet','sum_scrabe_shrinkage_bySet','sum_scrabe_dimentions_bySet','sum_scrabe_weight_bySet','sum_scrabe_dirty_bySet',
            'sum_scrabe_cloration_bySet','sum_scrabe_no_parts','number_scrab_by_item',
            'parts_patchsNumbers','gross_production','scrap_percent_by_item','Items_patchsNumbers','product_by_set_or_no',
            'parts_symbole','part_id','id_DayPartUnique','machine_type','factory','tall_mm','width_mm','deepth_mm','customer_name']

            daily_input= pd.read_excel(self.readfile,"input")
            #Block.show_yearly_report_itemsByMonths(self,year,month)
            #daily_input=cursor.fetchall()
            #data_analysis3=daily_input[cols_monthly]
            last_year=daily_input["year"].max()
            mold_analysis_bool4=daily_input[daily_input["year"]==last_year]
            #mold_analysis3=daily_input[mold_analysis_bool4]
            last_month=mold_analysis_bool4["month"].max()
            
            
            data_analysis3=daily_input[cols_monthly]
            data_analysis2=data_analysis3[data_analysis3["year"]==last_year]
            data_analysis1=data_analysis2[data_analysis2["month"]==last_month]

            year=last_year
            month=last_month
            
            print("_____________year_____",last_year,last_month)            
            #create list of items in monthes
            #list of items
            
            list_item=data_analysis1.groupby(["item_id"])["product_name"].unique()
            
            #list_item = {word: i for i, word in enumerate(items)}
            #list_item=pd.DataFrame()
            #list_item=list_item.append(items)
            list_item_size=pd.DataFrame(list_item,columns=["product_name"]).shape[0]

            #create  the index sheet:
            r = 2  # start at 4th row
            c = 1 # column 'a'
            for row in range(0,list_item_size):  
                rows = list_item.iloc[row]
                for item in rows:
                    ws_index.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 1
                r += 1
            #generate sheets:

            print("items",list_item.index)
            data_analysis=data_analysis1[data_analysis1["item_id"]==mold]

            #for creating data to sheets
            ws_index['g1'] = month            #day for xpscreport
            ws_index['i1'] = year    #month for xps report  
            
            #for i in list_item:
                #ws_copy=wb_formats.copy_worksheet(ws_input)
                #ws_copy.title=str(i)

            #data_analysis=data_analysis1[data_analysis1["item_id"]==254]


            
            #create spc for each mold#________________
            ws_data['g1'] = month            #day for xpscreport
            ws_data['i1'] = year    #month for xps report
            ws_data['c1']=data_analysis["product_name"].iloc[0]
#            ws_data['c1']=str(data_analysis["product_name"].head(1))

            ##analysis weight
            weight2=data_analysis[["day","machine_id","standard_dry_weight",
            "standard_dry_weight_from","standard_dry_weight_to","average_dry_weight"]]
            weight=weight2[weight2["average_dry_weight"].notnull()]
            
            weight_size=pd.DataFrame(weight,columns=["day"]).shape[0]
            r = 18  # start at 4th row
            c = 1 # column 'a'
            for row in range(0,weight_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
                rows = weight.iloc[row]
                for item in rows:
                    ws_data.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 1
                r += 1
            print("_______________test____________")
            print(weight)
            print("weight_size",weight_size)

            ##analysis ct
            ct2=data_analysis[["c_t_standard_per_second"]]
            ct2["c_t_standard_per_second_from"]=ct2["c_t_standard_per_second"]*0.95
            ct2["c_t_standard_per_second_to"]=ct2["c_t_standard_per_second"]*1.05
            ct2["c_t_actually"]=data_analysis["c_t_actually"]
            ct=ct2[ct2["c_t_actually"].notnull()]
            ct_size=pd.DataFrame(ct,columns=["day"]).shape[0]
            
            r = 18  # start at 18th row
            c = 8 # column 'h'
            for row in range(0,ct_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
                rows = ct.iloc[row]
                for item in rows:
                    ws_data.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 8
                r += 1
            ##analysis scrap
            scrap3=data_analysis[["day",'number_scrab_by_item','scrabe_standard','gross_production','scrap_percent_by_item']]
            scrap2=scrap3[scrap3["gross_production"].notnull()]
            scrap=scrap2[scrap2["gross_production"] > scrap2["number_scrab_by_item"]]
            
            #scrap=scrap2
            scrap_size=pd.DataFrame(scrap,columns=["day"]).shape[0]
            
            r = 18  # start at 4th row
            c = 13 # column 'n'
            
            for row in range(0,scrap_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
                rows = scrap.iloc[row]
                for item in rows:
                    ws_data.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 13
                r += 1

            #add charts
            #wb_formats.save(str(year)+"-"+str(month)+"MOLD"+str(mold)+" QC_SPC.xlsx")    
            
        ####add chart____________________
            #ws_data=wb_formats.copy_worksheet(ws_spc)
            
            #ws_data=wb_formats.get_sheet_by_name("spc Copy")
            #ws_data.title=str(mold)  
            # create data for plotting 
            chart_weight = LineChart()
            
            #values_weight = Reference(ws_data, min_col = 3, min_row = 18,  max_col = 6, max_row = 40) 
            x_weight = Reference(ws_data, min_col = 1, min_row = 18, max_col = 1, max_row = 40) 
            
            y_weight = Reference(ws_data, min_col = 4, min_row = 18, max_col = 6, max_row = 40) 
            
            #size_weight = Reference(ws_data, min_col = 1, min_row = 18
            # , max_row = 40)
            # create a 1st series of data
            #values_weight = Series(values = y_weight, xvalues = x_weight, zvalues = size_weight , title ="Weight")
            values_weight = Series(values = y_weight, xvalues = x_weight, title ="Weight")
            # Create object of BarChart class 
            
        
            # adding data to the Bar chart object 

            chart_weight.add_data(y_weight) 
            #chart_weight.add_data(x_weight) 
            chart_weight.series.append(values_weight)

            #chart_weight.add_data(values_weight) 
            
            # set the title of the chart 
            chart_weight.title = " dry weight "
            # set the title of the x-axis 
            chart_weight.x_axis.title = " X_AXIS "
            # set the title of the y-axis 
            chart_weight.y_axis.title = " Y_AXIS "
            ws_data.add_chart(chart_weight, "a2") 

            #_____
            #values_ct = Reference(ws_data, min_col = 8, min_row = 17,  max_col = 11, max_row = 40) 
            x_ct = Reference(ws_data, min_col = 9, min_row = 18, max_col = 10, max_row = 40) 
            y_ct = Reference(ws_data, min_col = 11, min_row = 18, max_col = 11, max_row = 40) 
            #size_ct = Reference(ws_data, min_col = 1, min_row = 18, max_row = 40)
            # create a 1st series of data
            values_ct = Series(values = y_ct, xvalues = x_ct, title ="CT")
            # add series data to the chart object
            
            chart_ct = LineChart()
            chart_ct.add_data(x_ct) 
            chart_ct.add_data(y_ct) 
            #chart_ct.series.append(values_ct)

            chart_ct.title = " ct-CHART "
            
            # set the title of the x-axis 
            chart_ct.x_axis.title = " X_AXIS "
            
            # set the title of the y-axis 
            chart_ct.y_axis.title = " Y_AXIS "
            ws_data.add_chart(chart_ct, "h2") 

            #____
            chart_scrap = LineChart()

            #values_scrab = Reference(ws_data, min_col = 13, min_row = 17, max_col = 17, max_row = 40) 
            x_scrab = Reference(ws_data, min_col = 13, min_row = 18, max_col = 11, max_row = 40) 
            y_scrab = Reference(ws_data, min_col = 17, min_row = 18, max_col = 17, max_row = 40) 
            z_scrab = Reference(ws_data, min_col = 15, min_row = 18, max_col = 15, max_row = 40) 
            #size = Reference(ws_data, min_col = 10, min_row = 17, max_row = 40)
            # create a 1st series of data
            #series = Series(values = y_scrab, xvalues = y_scrab, zvalues = size, title ="scrap")
            # add series data to the chart object
            series = Series(values = y_scrab, xvalues = y_scrab, zvalues = z_scrab, title ="scrap")
            #chart_scrap.series.append(series)
            chart_scrap.add_data(x_scrab)
            chart_scrap.add_data(y_scrab)
            chart_scrap.add_data(z_scrab)

            chart_scrap.title = " Scrap-CHART "

            # set the title of the x-axis 
            chart_scrap.x_axis.title = " X_AXIS "
    
            # set the title of the y-axis 
            chart_scrap.y_axis.title = " Y_AXIS "
            ws_data.add_chart(chart_scrap, "o2") 
            # add chart to the sheet 
            # the top-left corner of a chart 
            # is anchored to cell E2 . 
            
                    
                
            #create t
            #for s in wb_formats:
            #values = Reference(scrap)
            #chart = BarChart()
            #chart.add_data(values)
            #ws = wb_formats[ws_spc]
            #ws.add_chart(chart, "B5")
    
        
            wb_formats.save(save_name)