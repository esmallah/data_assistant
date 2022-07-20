'''this module for the data was emported from database analysis'''
from calendar import month
import csv
from statistics import mean
import pandas as pd
import openpyxl as xl
from openpyxl import load_workbook

import os
import numpy as np
import glob

#import database_postgrsql as database
from apps.analysis.db_reports import Block,cursor,conn,Material
from random import randint,seed

#master data
    #for columns from ite_ master for molds ( the one product from one moldes)
#columns_machines=['id','scrabe_standard','machine_type','place' ,'low_size','high_size']

columns_molds=['mold_id','ORG_CODE','ORG_NAME','Ctegory','UOM','machie_size','No_on_Set',
'set','No_on_Set',"sub_category",'mold_name','status','c_t_standard_per_second','c_t_standard_per_second_from',
'c_t_standard_per_second_to',"view_molds",
'gap_mm','injekors_numbers','drummers_number','customer_id','playback','gap_tolerance',
'gap_mm_from','gap_mm_to','machine_parameter_id','mold_start_date','mold_expire_date',"customer_name","company_of_customer",
"customer_proberty_name","customer_proberty_sequence","customer_asset_code","Customer_Product_Group"]
 
#for columns to more than one products form the same mold
columns_parts=['part_id','product_code','product_name_by_parts','Weight_kg','standard_rate_hour','highlite',
    'standard_dry_weight','standard_dry_weight_from','standard_dry_weight_to','positive_weight','negative_weight',
    'sub_category','mold_id','product_parts','item_id','product_name','item_classification_customers','item_code_customers','item_classification_customers'
    ,"view_items","view_molds",
    'view_parts','density','row_material_typeA','row_material_typeB',
    'tall_mm','tall_positive_tolerance','tall_negative_tolerance','width_mm','width_positive_tolerance','width_negative_tolerance',
    'sicness_mm','sicness_positive_tolerance','sicness_negative_tolerance','id_printed_specification','spec_folder_no',
    'page_volum_x','page_volum_y','page_volum_z','volume','sitotb_color','sitotb_set','silotib_meter_reels',
    'silotib_outside_meter','package_page','number_bacage','page_size_x','page_size_y','page_colore','pages_kgm_set','pages_kgm',
    'kg_after_add12percent','kg_after_add8_5percent','id_modification','notes','silotib_inside_meter']

#prepare data entering tables
    #3collecting tables ct,weight and scrap(yt_quality)
#columns for data entry only
columns_quality=['year',
    'month',
    'day',
    'machine_id',
    'item_id',
    'mold_id',
    'set',
    'no_on_set',
    'shift1_wet_weight1',
    'shift1_wet_weight2',
    'shift1_wet_weight3',
    'shift1_wet_weight4',
    'shift1_wet_weight5'
    ,'shift1_dry_weight1','shift1_dry_weight2','shift1_dry_weight3','shift1_dry_weight4','shift1_dry_weight5',
    'shift1_c_t1','shift1_c_t2',
    'shift2_wet_weight1','shift2_wet_weight2','shift2_wet_weight3','shift2_wet_weight4','shift2_wet_weight5',
    'shift2_dry_weight1','shift2_dry_weight2','shift2_dry_weight3','shift2_dry_weight4','shift2_dry_weight5',
    'shift2_c_t1','shift2_c_t2'

    ,'shift1_production_cards'
    ,'shift1_prod_page','shift1_proper_production','shift1_scrabe_shortage','shift1_scrabe_roll','shift1_scrabe_broken',
    'shift1_scrabe_curve','shift1_scrabe_shrinkage','shift1_scrabe_dimentions','shift1_scrabe_weight','shift1_scrabe_dirty'
    ,'shift1_scrabe_cloration','shift2_production_cards','shift2_prod_page','shift2_proper_production','shift2_scrabe_shortage',
    'shift2_scrabe_roll','shift2_scrabe_broken','shift2_scrabe_curve','shift2_scrabe_shrinkage','shift2_scrabe_dimentions'
    ,'shift2_scrabe_weight','shift2_scrabe_dirty','shift2_scrabe_cloration',

    'standard_dry_weight','standard_dry_weight_from','standard_dry_weight_to','c_t_standard_per_second','standard_rate_hour','scrabe_standard',
    'customer_name',
    'item_classification_customers',
    'item_code_customers',
    ]
    
    #the previous columns separate to 
        #1 cycle time table
columns_cycle_time=['year','month','day','machine_id','rat_actually','c_t_deviation','shift1_c_t1','shift1_c_t2'
                ,'shift2_c_t1','shift2_c_t2','c_t_actually','mold_id','part_id','item_id','number_day_use','factory','id_DayPartUnique'
                ,'parts_patchsNumbers','Items_patchsNumbers','bachStartDate','bachEndDate']

        #2 quality inspections scrap and weights
columns_QCinspection=['year','month','day','machine_id','item_id','number_day_use','mold_id','product_parts',
    'shift1_wet_weight1','shift1_wet_weight2','shift1_wet_weight3','shift1_wet_weight4','shift1_wet_weight5',
    'shift1_dry_weight1',    'shift1_dry_weight2','shift1_dry_weight3','shift1_dry_weight4','shift1_dry_weight5','shift1_ct1','shift1_ct2',
    'shift2_wet_weight1','shift2_wet_weight2','shift2_wet_weight3','shift2_wet_weight4','shift2_wet_weight5',
    'shift2_dry_weight1','shift2_dry_weight2','shift2_dry_weight3','shift2_dry_weight4','shift2_dry_weight5',

    'shift2_ct1','shift2_ct2'

    ,'average_wet_weight','average_dry_weight','rat_actually','shift1_production_cards'
    ,'shift1_prod_page','shift1_proper_production','shift1_scrabe_shortage','shift1_scrabe_roll','shift1_scrabe_broken',
    'shift1_scrabe_curve','shift1_scrabe_shrinkage','shift1_scrabe_dimentions','shift1_scrabe_weight','shift1_scrabe_dirty'
    ,'shift1_scrabe_cloration','shift1_scrabe_no_parts'
    

    ,'shift2_production_cards','shift2_prod_page','shift2_proper_production','shift2_scrabe_shortage',
    'shift2_scrabe_roll','shift2_scrabe_broken','shift2_scrabe_curve','shift2_scrabe_shrinkage','shift2_scrabe_dimentions'
    ,'shift2_scrabe_weight','shift2_scrabe_dirty','shift2_scrabe_cloration'    
    ,'bachStartDate',
'parts_patchsNumbers','Items_patchsNumbers','bachStartDate','bachEndDate']

            #the previeuse table separate to
                    #1weights table
columns_weight=['year','month','day','machine_id','item_id','number_day_use','mold_id','product_parts',
'shift1_dry_weight1','shift1_dry_weight2','shift1_dry_weight3','shift1_dry_weight4'
,'shift1_dry_weight5','shift2_dry_weight1','shift2_dry_weight2','shift2_dry_weight3','shift2_dry_weight4',
'shift2_dry_weight5','average_wet_weight','average_dry_weight','dryweight_deviation_validation'
,'parts_patchsNumbers','Items_patchsNumbers','bachStartDate','bachEndDate','standard_dry_weight_from','standard_dry_weight_to','standard_dry_weight',]

                    #2scrab table
columns_scrap=['year','month','day','machine_id','item_id','mold_id','product_parts','part_id','shift1_production_cards'
,'shift1_prod_page','shift1_proper_production','shift1_scrabe_shortage','shift1_scrabe_roll','shift1_scrabe_broken',
'shift1_scrabe_curve','shift1_scrabe_shrinkage','shift1_scrabe_dimentions','shift1_scrabe_weight','shift1_scrabe_dirty'
,'shift1_scrabe_cloration','shift1_scrabe_no_parts','shift1_scrabe_no_item','shift1_all_production'
,'shift2_production_cards','shift2_prod_page','shift2_proper_production','shift2_scrabe_shortage',
'shift2_scrabe_roll','shift2_scrabe_broken','shift2_scrabe_curve','shift2_scrabe_shrinkage','shift2_scrabe_dimentions'
,'shift2_scrabe_weight','shift2_scrabe_dirty','shift2_scrabe_cloration','shift2_scrabe_no_parts','shift2_scrabe_no_item'
,'shift2_all_production','sum_scrabe_shortage_bySet','sum_scrabe_roll_bySet','sum_scrabe_broken_bySet'
,'sum_scrabe_curve_bySet','sum_scrabe_shrinkage_bySet','sum_scrabe_dimentions_bySet','sum_scrabe_weight_bySet','sum_scrabe_dirty_bySet',
'sum_scrabe_cloration_bySet','sum_scrabe_no_parts','number_scrab_by_item',
'gross_production','scrap_percent_by_item','factory','id_DayPartUnique','Items_patchsNumbers'
,'parts_patchsNumbers','bachStartDate','bachEndDate']


column_monthly_report=[
                'year',
				'month',
				'day',
                'mold_id',
				'item_id',
                "product_name",
                "product_code",
                
                "standard_dry_weight",
                "standard_dry_weight_from",
                "standard_dry_weight_to",
                #"c_t_deviation",
                #'rat_validation',
                                
                
                "standard_dry_weight",
                "average_dry_weight",
                
				'average_wet_weight',
                'wet_average_percent',
                "standard_rate_hour",
                "c_t_standard_per_second",
                "rat_actually",
                "c_t_actually",

                'sum_scrabe_shortage_bySet',
                'sum_scrabe_roll',
                'sum_scrabe_broken',
                'sum_scrabe_curve',
                'sum_scrabe_shrinkage',
                'sum_scrabe_dimentions',
                'sum_scrabe_weight',
                'sum_scrabe_dirty_bySet',
                'sum_scrabe_cloration',
                "scrabe_standard",
                
                'number_scrab_by_item',
                'gross_production',
                
                "standard_scrap_weight_kg",
                "standard_production_weight_kg",
                "scrap_weight_kg",
                "production_weight_kg",
                "HoursScrap",
                "mold_avalibility",
				'number_day_use',
				'customer_name',
                'item_classification_customers',
                'item_code_customers',
                "scrap_percent_by_item"

]
#error __________________
#must be reduce column by ceate calculating columns

col_rename={"product_name": "اسم المنتج","product_name":"اسم المنتج بالاجزاء"
        ,"product_code":"كود المنتج","product_parts":"اجزاء المنتج","machine_id":'رقم الماكينة',"number_day_use":"عدد ايام التشغيل"
        ,"standard_dry_weight_from":"مواصفة الوزن الجاف من","standard_dry_weight_to":"مواصفة الوزن الجاف إلي",
        "weight_under_validation":"الاوزان الاقل من المواصفة","weight_above_validation":"الاوزان الاعلي من المواصفة",
        #"c_t_deviation":"الفارق بين المعدل المعيارى والفعلى","rat_standard_deviation":'الانحراف المعياري لمعدل الانتاج',
        
        "sum_scrabe_no_parts":"عدد التوالف بالقطع","number_scrab_by_item":"عدد التوالف بالطقم",
        "gross_production":"إجمالي الانتاج بالقطعة","scrap_percent_by_item":"نسبة الاسكراب بالطقم",
        "product_groub":"اسم المنتج المجمع","standard_dry_weight":"الوزن الجاف المعياري","average_dry_weight":"متوسط الوزن الجاف الفعلي"
        ,"standard_rate_hour":"المعدل المعياري بالساعة","rat_actually":"متوسط المعدل الفعلي بالساعة"
        ,"c_t_standard_per_second":"زمن الدورة المعياري بالثانية","c_t_actually":"متوسط زمن الدورة الفعلي بالثانية"
        ,"scrabe_standard":"معياري الاسكراب","parts_patchsNumbers":"ارقام الباتشات بالجزء","scrap_percent_by_item":"نسبة التوالف بالقطعة"
        ,"Items_patchsNumbers":"ارقام الباتشات بالاصناف","scrabe_standard":"معياري التوالف","part_id":"رقم الجزء",
        "mold_id":"رقم الاسطمبة","item_id":"رقم المنتج" }      # for recaull module to change english name to arabic

columns_weight=["day","year","month","machine_id","product_name","product_code","standard_dry_weight",
"standard_dry_weight_from","standard_dry_weight_to",'shift1_dry_weight1','shift1_dry_weight2','shift1_dry_weight3',
        'shift1_dry_weight4','shift1_dry_weight5','shift2_dry_weight1','shift2_dry_weight2',
        'shift2_dry_weight3','shift2_dry_weight4','shift2_dry_weight5',"average_dry_weight","average_wet_weight"]

columns_machine=["day","year","month","machine_id","product_name","product_code","scrabe_standard",
"sum_scrabe_no_parts","number_scrab_by_item","gross_production",'scrap_percent_by_item'
,"number_day_use","item_id","machine_type"]

columns_cycle_time=["day","year","month","machine_id","mold_name","product_name","product_code","set","standard_rate_hour"
,"c_t_standard_per_second",'shift1_c_t1','shift1_c_t2','shift2_c_t1','shift2_c_t2',"rat_actually","c_t_actually","number_day_use","mold_id",]

class Unique():
    def __init__(self,folder,readfile,readsheet,column1,column2,writefile,sheetwriter):
        #self.folder=folder
        self.folder=folder
        self.readfile=readfile
        self.readsheet=readsheet
        self.column1=column1
        self.column2=column2
        self.writefile=writefile
        self.sheetwriter=sheetwriter
    
    def unique_list(self):
        ''' to get unique data in any excel sheet'''
        os.chdir(self.folder)
        reader=pd.read_excel(self.readfile,self.readsheet)
        data=reader[[self.column1,self.column2]]
        
        unique_data=data.groupby(self.column1)[self.column2].count()
        writer = pd.ExcelWriter(self.writefile)
        
        unique_data.to_excel(writer,self.sheetwriter)
        writer.save()
    def item_master(self):
        '''for create item master from table item_master.exe'''
        os.chdir(self.folder)
        reader=pd.read_excel(self.readfile,self.readsheet)
        data=reader[columns_molds]
        unique_data=data.groupby(columns_molds)[self.column2].count()

        writer = pd.ExcelWriter(self.writefile)
        print(writer)
        unique_data.to_excel(writer,self.sheetwriter)
        writer.save()
    def report_analysis(self):
        os.chdir(self.folder)
        
        reader3=pd.read_excel(self.readfile,self.readsheet)
        last_year=reader3["year"].max()
        reader2_bool=reader3["year"]==last_year
        reader2=reader3[reader2_bool]
        last_month=reader2["month"].max()
        reader2_bool=reader2["month"]==self.sheetwriter
        reader=reader2[reader2_bool]
        unique_data=reader.groupby(self.column1)[self.column2].count()
        analysis_data=reader.groupby(self.column1)[self.column2].describe()
        
        #weights analsysis
        molds_weight_bool2=reader["average_dry_weight"]<reader["standard_dry_weight_from"]
        molds_weight3=reader[molds_weight_bool2]
        molds_weight_bool=molds_weight3["average_dry_weight"]>molds_weight3["standard_dry_weight_to"]
        molds_weight2=molds_weight3[molds_weight_bool]
        molds_weight_ncr=molds_weight2.groupby("mold_name")["average_dry_weight"].mean()

        #scrap analsysis
        reader["scrap_percent_by_set_all"]=reader['number_scrab_by_item']/reader['gross_production']
        molds_scrap_bool=reader["scrabe_standard"]<reader["scrap_percent_by_set_all"]
        molds_scrap2=reader[molds_scrap_bool]
        molds_scrap_ncr=molds_scrap2.groupby("mold_name")["gross_production","number_scrab_by_item","sum_scrabe_no_parts"].sum()

        #ct analsysis
        molds_ncr_rat_bool=reader["rat_actually"]<reader["standard_rate_hour"]
        molds_ncr_rat2=reader[molds_ncr_rat_bool]
        molds_ncr_rat=reader.groupby("mold_name")["rat_actually","c_t_actually"].mean()

        writer = pd.ExcelWriter(self.writefile)
        print(writer)
        unique_data.to_excel(writer,"Sheet1")
        analysis_data.to_excel(writer,"analysis")
        molds_ncr_rat.to_excel(writer,"molds_ncr_rat")
        molds_weight_ncr.to_excel(writer,"molds_weight_ncr")
        molds_scrap_ncr.to_excel(writer,"molds_scrap_ncr")
        writer.save()
    def returns_report(self):
        os.chdir(self.folder)
        reader=pd.read_excel(self.readfile,self.readsheet)
        return_day_bool=reader["month"]==self.column1
        return_day=reader[return_day_bool]
        unique_data=return_day.groupby(self.column2).sum()
        writer = pd.ExcelWriter(self.writefile)
        print(writer)
        unique_data.to_excel(writer,self.sheetwriter)
        writer.save()
    
    def return_crosstab(self,year,**customer):
        print("_____________return report________________")
        '''for get dtata form qc return'''
        os.chdir(self.folder)

        #for development this function in this erea>> git data from database for mix warehouse data with quality data
        reader_data=pd.read_excel(self.readfile,self.readsheet)
        last_year=reader_data["year"].max()
        reader2_bool=reader_data["year"]==last_year
        reader2=reader_data[reader2_bool]
        reader_filter_year=reader2[reader2["year"]==last_year]
        
        reader=reader_filter_year
        #reader=reader_filter_year[reader_filter_year["customer_name"]=="customer"]
        #reader=reader_filter_year[reader_filter_year["product_type"]=="xps"]
        
        df = pd.DataFrame(reader, columns = ['customer_name','customer_name_category', 'product_name', 'product_code', 'year',
         'month', 'date',"gross_quantity","return_quantity","scrap","return_reason",'inspection_results_proper','inspection_results_low_grade',
         'product_type','reason_category','lenth','width','denesity'])
        
        sheets=['customer_name', 'product_name','product_type']
        print("________________test_______________")
        print(df.reason_category)
        #for development this function in this erea>> git output by all delivers to customer and quantity return

        writer = pd.ExcelWriter(self.writefile)
        
        #to return quantity
        for f in sheets:
            reason_data=pd.crosstab([df[f],df.reason_category],[df.year,df.month],
            df.return_quantity,aggfunc='sum',margins=False)
            reason_data.to_excel(writer,f+"_reasons")        
            #to scrap sheet
            scrap=pd.crosstab([df[f],df.reason_category],[df.year,df.month],
            df.scrap,aggfunc='sum',margins=False)
        for f in sheets:
            unique_data=pd.crosstab([df[f]],[df.year,df.month],df.return_quantity,aggfunc='sum',margins=False)
            unique_data.to_excel(writer,f)     
        for f in sheets:
            scrap=pd.crosstab([df[f]],[df.year,df.month],df.scrap,aggfunc='sum',margins=False)
            scrap.to_excel(writer,f+"_scrap")     

        #to input quantity
        reader.to_excel(writer,"input")        
        writer.save()

class Select():
    
    """this class provide  work books and sheet names as input """
    def __init__(self,folder,readfile1,sheet1,year,month,writefile,writesheet):
        self.folder=folder
        self.readfile1=readfile1
        self.sheet1=sheet1
        self.year=year
        self.month=month
        self.writefile=writefile
        self.writesheet=writesheet
    def load_data(self,sql_query):
        from apps import cursor
        sql_query
        get_data2=cursor.fetchall()
        #__________________________________________________________________        
        column_names = [desc[0] for desc in cursor.description]
        get_data = pd.DataFrame(get_data2,columns=column_names)
        
        data=get_data[columns_quality]

        #brebare data
        
        dry_weight_cl=[
        'shift1_dry_weight1','shift1_dry_weight2',
        'shift1_dry_weight3','shift1_dry_weight4','shift1_dry_weight5',
        'shift2_dry_weight1','shift2_dry_weight2','shift2_dry_weight3',
        'shift2_dry_weight4','shift2_dry_weight5',
        ]
        
        data['average_dry_weight']=data[dry_weight_cl].mean(axis=1)
        wet_weight_cl=[
        'shift1_wet_weight1',
        'shift1_wet_weight2',
        'shift1_wet_weight3',
        'shift1_wet_weight4',
        'shift1_wet_weight5',
        'shift2_wet_weight1',
        'shift2_wet_weight2',
        'shift2_wet_weight3',
        'shift2_wet_weight4',
        'shift2_wet_weight5']
        data['average_wet_weight']=data[wet_weight_cl].mean(axis=1)
        
        cycletime_cl=['shift1_c_t1','shift1_c_t2','shift2_c_t1','shift2_c_t2']
    
        data['c_t_actually']=data[cycletime_cl].mean(axis=1)
        
        data['rat_actually']=(3600 / data['c_t_actually']) * data['set'].fillna(0).astype(int)
        
        data['c_t_actually']=data['c_t_actually']
        
        scrab_shift1_cl=[
        'shift1_scrabe_shortage',
        'shift1_scrabe_roll','shift1_scrabe_broken','shift1_scrabe_curve','shift1_scrabe_shrinkage','shift1_scrabe_dimentions'
        ,'shift1_scrabe_weight','shift1_scrabe_dirty','shift1_scrabe_cloration',   ]
        scrab_shift2_cl=[
        'shift2_scrabe_shortage',
        'shift2_scrabe_roll','shift2_scrabe_broken','shift2_scrabe_curve','shift2_scrabe_shrinkage','shift2_scrabe_dimentions'
        ,'shift2_scrabe_weight','shift2_scrabe_dirty','shift2_scrabe_cloration', ]

        data['shift1_scrabe_no_parts']=data[scrab_shift1_cl].fillna(0).astype(int).sum(axis=1)
        data['shift2_scrabe_no_parts']=data[scrab_shift2_cl].fillna(0).astype(int).sum(axis=1)
        data['sum_scrabe_no_parts']=data[['shift1_scrabe_no_parts','shift2_scrabe_no_parts',]].sum(axis=1)
        data['shift1_scrabe_no_item']=data['shift1_scrabe_no_parts'].fillna(0).astype(int)/ data['no_on_set']
        data['shift2_scrabe_no_item']=data['shift2_scrabe_no_parts'].fillna(0).astype(int)/ data['no_on_set']
        data['shift1_all_production']=data[['shift1_proper_production','shift1_scrabe_no_item']].fillna(0).astype(int).sum(axis=1)
        data['shift2_all_production']=data[['shift2_proper_production','shift2_scrabe_no_item']].fillna(0).astype(int).sum(axis=1)
        data['number_scrab_by_item']=data[['shift1_scrabe_no_item','shift2_scrabe_no_item']].sum(axis=1)
        data['gross_production']=data[['shift1_all_production','shift2_all_production']].sum(axis=1)
        data['sum_scrabe_shortage_bySet']=data[['shift1_scrabe_shortage','shift2_scrabe_shortage']].sum(axis=1)
        data['sum_scrabe_roll']=data[['shift1_scrabe_roll','shift2_scrabe_roll']].sum(axis=1)
        data['sum_scrabe_broken']=data[['shift1_scrabe_broken','shift2_scrabe_broken']].sum(axis=1)
        data['sum_scrabe_curve']=data[['shift1_scrabe_curve','shift2_scrabe_curve']].sum(axis=1)
        data['sum_scrabe_shrinkage']=data[['shift1_scrabe_shrinkage','shift2_scrabe_shrinkage']].sum(axis=1)
        data['sum_scrabe_dimentions']=data[['shift1_scrabe_dimentions','shift2_scrabe_dimentions']].sum(axis=1)
        data['sum_scrabe_weight']=data[['shift1_scrabe_weight','shift2_scrabe_weight']].sum(axis=1)
        data['sum_scrabe_dirty_bySet']=data[['shift1_scrabe_dirty','shift2_scrabe_dirty']].sum(axis=1)
        data['sum_scrabe_cloration']=data[['shift1_scrabe_cloration','shift2_scrabe_cloration']].sum(axis=1)
        data['HoursScrap']=data['number_scrab_by_item']/data['rat_actually'] #number hourse of scrap*/,
        data['number_day_use']=data['average_dry_weight'].count()
        
        data['mold_avalibility']=data['gross_production'] * (data['number_day_use']/22) * data['standard_rate_hour'].fillna(0).astype(int)#as  /*avalibility bercent in 22 work hours */
        #notece that
        data['standard_scrap_weight_kg']= data['number_scrab_by_item']/data['standard_dry_weight'].fillna(0).astype(int)
        data['standard_production_weight_kg']=data['gross_production']/data['standard_dry_weight'].fillna(0).astype(int)
        #data['scrap_weight_kg'=data.number_scrab_by_item.sum()/data.average_dry_weight.mean()
        #data.production_weight_kg= data.gross_production.sum()/data.average_dry_weight.mean()
        data['production_weight_kg']= data['gross_production']/data['average_dry_weight']
        data['scrap_weight_kg']=data['number_scrab_by_item']/data['average_dry_weight']
                
        get_data['average_dry_weight']=data['average_dry_weight']
        get_data['average_wet_weight']=data['average_wet_weight']
        data['wet_average_percent']=data['average_wet_weight']-data['average_wet_weight']-data['standard_dry_weight'].fillna(0).astype(int)/data['standard_dry_weight'].fillna(0).astype(int)
        get_data['standard_dry_weight']=data['standard_dry_weight'].fillna(0).astype(int)
        get_data['standard_dry_weight_from']=data['standard_dry_weight_from'].fillna(0).astype(int)
        get_data['standard_dry_weight_to']=data['standard_dry_weight_to'].fillna(0).astype(int)
        get_data['scrabe_standard']=data['scrabe_standard'].fillna(0).astype(float)
        get_data['c_t_standard_per_second']=data['c_t_standard_per_second'].fillna(0).astype(int)
        #الوزن المبلل - معياري الوزن الجاف للصنف/معياري الوزن الجاف للصنف
        get_data['standard_rate_hour']=data['standard_rate_hour'].fillna(0).astype(int)
        get_data['rat_actually']=data['rat_actually']
        get_data['c_t_actually']=data['c_t_actually']
        ##### end the query and start anasist will put in sheet by order
        
        

        get_data['shift1_scrabe_no_parts']=data['shift1_scrabe_no_parts']
        
    
        get_data['shift1_scrabe_no_item']=data['shift1_scrabe_no_item']
        get_data['shift1_all_production']=data['shift1_all_production']  
        get_data['shift2_scrabe_no_parts']=data['shift2_scrabe_no_parts']
        get_data['shift2_scrabe_no_item']=data['shift2_scrabe_no_item']
        get_data['shift2_all_production']=data['shift2_all_production']

        get_data['sum_scrabe_shortage_bySet']=(data['sum_scrabe_shortage_bySet'].fillna(0).astype(int)/ data['no_on_set']).fillna(0).astype(int)
        get_data['sum_scrabe_roll']=(data['sum_scrabe_roll'].fillna(0).astype(int)/ data['no_on_set']).fillna(0).astype(int)
        get_data['sum_scrabe_broken']=(data['sum_scrabe_broken'].fillna(0).astype(int)/ data['no_on_set']).fillna(0).astype(int)
        get_data['sum_scrabe_curve']=(data['sum_scrabe_curve'].fillna(0).astype(int)/ data['no_on_set']).fillna(0).astype(int)
        get_data['sum_scrabe_shrinkage']=(data['sum_scrabe_shrinkage'].fillna(0).astype(int)/ data['no_on_set']).fillna(0).astype(int)
        get_data['sum_scrabe_dimentions']=(data['sum_scrabe_dimentions'].fillna(0).astype(int)/ data['no_on_set']).fillna(0).astype(int)
        get_data['sum_scrabe_weight']=(data['sum_scrabe_weight'].fillna(0).astype(int)/ data['no_on_set']).fillna(0).astype(int)
        get_data['sum_scrabe_dirty_bySet']=(data['sum_scrabe_dirty_bySet'].fillna(0).astype(int)/ data['no_on_set']).fillna(0).astype(int)
        get_data['sum_scrabe_cloration']=(data['sum_scrabe_cloration'].fillna(0).astype(int)/ data['no_on_set']).fillna(0).astype(int)
        get_data['sum_scrabe_no_parts']=data['sum_scrabe_no_parts']
        get_data['number_scrab_by_item']=data['number_scrab_by_item']    
        get_data['gross_production']=data['gross_production']    
        get_data['scrap_percent_by_item']=data['number_scrab_by_item']/data['gross_production']
        get_data['standard_scrap_weight_kg']= data['standard_scrap_weight_kg']
        get_data['standard_production_weight_kg']=data['standard_production_weight_kg']
        get_data['scrap_weight_kg']=data['scrap_weight_kg']
        get_data['production_weight_kg']=data['production_weight_kg']
        get_data["HoursScrap"]=data['HoursScrap']
        get_data['mold_avalibility']=data['mold_avalibility']
        get_data['number_day_use']=data['number_day_use'].fillna(0).astype(int)
        get_data['wet_average_percent']=data['wet_average_percent']
        get_data['customer_name']=data['customer_name']
        get_data['item_classification_customers']=data['item_classification_customers']
        get_data['item_code_customers']=data['item_code_customers']


        #to convert decimel
        decimals = pd.Series([1,1,1,1,1,1, 1,0,0,0], index=[
            'wet_average_percent',
            'production_weight_kg',
            'standard_production_weight_kg',
            
            'average_wet_weight',
            'average_dry_weight',
            'scrap_weight_kg',
            'mold_avalibility',
            'rat_actually'
            ,'c_t_actually',
            'HoursScrap'])
        get_data.round(decimals)    
        return get_data

    def select_data(self,year,month,day,isday=True,monthly=True,yearly=True,masterData=True,quality_records=True):
        print("select data starts")  
        '''to convert excel file to csv for entering to database
        already day is true for select day by day
        isday=False , its mean the selection for month
        '''
        os.chdir(self.folder)   # for work in the same folder

        #filter master data
        if masterData:
            master_data2=pd.read_excel(self.readfile1,"items_spec")  
            #        #upload  list without parts out of sevice

            #master_data=master_data2[master_data2["status"]==1]
            master_data=master_data2
            #upload moldss list 
            molds2=master_data[columns_molds]
            molds_bool=molds2["view_molds"].notnull()
            molds_list=molds2[molds_bool]
            #upload parts list 
            products_list=master_data[columns_parts]
            infr_data=pd.read_excel(self.readfile1,"machines_master")
        
        #filter inspection data
        daily_data3=pd.read_excel(self.readfile1,"input")
        material_data3=pd.read_excel(self.readfile1,"material")
        #filter the time
        last_year=daily_data3["year"].max()
        

        daily_data2=daily_data3[daily_data3["year"]==last_year]
        daily_data_material2=material_data3[material_data3["year"]==last_year]
        last_month=daily_data2["month"].max()

        daily_data1=daily_data2[daily_data2["month"]==int(month)]
        daily_data_material1=daily_data_material2[daily_data_material2["month"]==int(month)]
        last_day=daily_data1["day"].max()
        
        #to select periond of data
        if isday:
            daily_analysis=daily_data1[daily_data1["day"]==int(day)]
            daily_analysis_materia=daily_data_material1[daily_data_material1["day"]==int(day)]
        elif monthly:
            daily_analysis=daily_data1
            daily_analysis_materia=daily_data_material1
        elif yearly:
            daily_analysis=daily_data2
            daily_analysis_materia=daily_data_material2
        else:    
            daily_analysis=daily_data3
            daily_analysis_materia=material_data3
        #daily_analysis_materia=material_data3      #for get all rows###__________
        #daily_analysis=daily_data3                 #for get all rows###__________
        #filter master data
            #filter weights and scrap table
        productions_isnpection=daily_analysis[columns_quality]
         #filter ct tables
        molds_rate3=daily_analysis[columns_cycle_time]
        
        molds_rate_bool=molds_rate3["c_t_actually"].notnull()
        molds_rate2=molds_rate3[molds_rate_bool]
            #filter weight tables
        dry_weight3=daily_analysis[columns_weight]
        dry_weight2_bool=dry_weight3["average_dry_weight"].notnull()
        dry_weight2=dry_weight3[dry_weight2_bool]
            #filter scrab tables
        scrap3=daily_analysis[columns_scrap]
        scrap2_bool=scrap3["gross_production"].notnull()
        scrap2=scrap3[scrap2_bool]
        print("finishing the filtration and we will start to output now")  
        writer = pd.ExcelWriter("input_to_csv.xlsx")
        if masterData:
            molds_list.to_excel(writer,"molds_list", index=False)
            products_list.to_excel(writer,"parts_list", index=False)
            infr_data.to_excel(writer,"machines_list", index=False)
        else:
            productions_isnpection.to_excel(writer,"quality_records", index=False)        
            daily_analysis_materia.to_excel(writer,"materials", index=False)        
            dry_weight2.to_excel(writer,"weight_input", index=False)
            molds_rate2.to_excel(writer,"ct_input", index=False)        
            dry_weight2.to_excel(writer,"weight_input", index=False)
            scrap2.to_excel(writer,"scrap_input", index=False)
        writer.save()    

        print ("completed catch data  for ")
        print("year",daily_analysis["year"].unique())
        print("month",daily_analysis["month"].unique())
        print("for the days")
        print(daily_analysis["day"].unique())
        writer.save()
        
    def convert_csv(self,material=True,masterData=True,quality_records=True):
        os.chdir(self.folder)   # for work in the same folder
        '''to convert excel file to csv for entering to database'''
        
        #file_identifier = "*.XLS"
        #THIS_FOLDER = os.path.abspath(self.folder)
        df=pd.read_excel("input_to_csv.xlsx")
        if quality_records:
            pd.read_excel("input_to_csv.xlsx", "quality_records").to_csv("quality_records.csv", index=False)#input_to_database.csv
            pd.read_excel("input_to_csv.xlsx", "materials").to_csv("materials.csv", index=False)#input_to_database.csv

        if material:
            df=pd.read_excel("materials.xlsx")
            pd.read_excel("materials.xlsx", "Sheet1").to_csv("materials.csv", index=False)#input_to_database.csv
        if masterData:
            pd.read_excel("input_to_csv.xlsx", "machines_list").to_csv("machines.csv", index=False)#input_to_database.csv
            pd.read_excel("input_to_csv.xlsx", "molds_list").to_csv("molds_list.csv", index=False)#input_to_database.csv
            pd.read_excel("input_to_csv.xlsx", "parts_list").to_csv("parts_list.csv", index=False)#input_to_database.csv
        else:  
            pd.read_excel("input_to_csv.xlsx", "quality_records").to_csv("quality_records.csv", index=False)#input_to_database.csv
            pd.read_excel("input_to_csv.xlsx", "materials").to_csv("materials.csv", index=False)#input_to_database.csv
#_________________ to copy rang to another
    #File to be copied
        print("data is ready for upload to data base")
        #print(sheetname)        
#_________________ to copy rang to another
    #File to be copied
    def import_database(self):
    #    os.chdir(self.folder)
    
     #   daily_analysis=pd.read_csv("quality_records.csv")
      #  productions_isnpection=daily_analysis[columns_quality]
       # cur=database.cursor
        #table="yt_quality"
        

        Block.get_daily_dataentry_items(self)
        rows=cursor.fetchall()
        print(rows["month"])
        f = open('D:\work\contact_group\Contact records\QC quality control\Foam\qc_molds\database\quality_records.csv')
        cur.copy_from(f, table, sep=',')
        f.close()
    
        conn.commit()

    
    def export_report_mothly(self,writerFile,year,month,day,to_day,*args,monthly=True):
        '''
        to get daily report and monthly repots ended by QC_molds_daily_archive_v3
        '''
        from apps import Block,cursor,conn

        os.chdir(self.folder)
        wb = xl.load_workbook(self.readfile1)
        #__________________________________________________________________        

        sql_query=Block.get_daily_dataentry_items(self,year,month,day)
        get_data = self.load_data(sql_query)        

        list_item_size=get_data.shape[0]
        #get_data.shift1_all_production=sum(get_data.shift1_scrabe_shortage,get_data.shift1_scrabe_roll,get_data.shift1_scrabe_broken, get_data.shift1_scrabe_curve,  get_data.shift1_scrabe_shrinkage,get_data.shift1_scrabe_dimentions,get_data.shift1_scrabe_weight,get_data.shift1_scrabe_dirty)
        
        print('______________get data________________________',get_data['average_dry_weight'])
        #__________________________________________________________________
        #daily input

        ws1=wb["input_daily"]
        #create  the index sheet:
        r = 4  # start at 4th row
        c = 1 # column 'a'
        for row in range(0,list_item_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
            rows = get_data.iloc[row]
            for item in rows:
                ws1.cell(row=r, column=c).value = item
                c += 1 # Column 'd'
            c = 1
            r += 1   
        
        #material by silo
        ws_bach=wb["material_daily"]
        Material.material_bySilo_daily(self,year,month,day,to_day)
        get_data=cursor.fetchall()
        rows=get_data      
        r = 4  # start at fourd row
        c = 1 # column 'a'
        for row in rows:
            #print(row)
            for item in row:
                ws_bach.cell(row=r, column=c).value = item
                c += 1 # Column 'b'
            c = 1
            r += 1
        #Bache input
        ws_bach=wb["batches"]
        Block.show_monthly_Baches(self,year,month)
        
#        sql_query=Block.show_monthly_Baches(self,year,month)
 #       get_data = self.load_data(sql_query)        

        get_data=cursor.fetchall()
        rows=get_data
        #rows = get_data[columns_quality]
        
        r = 4  # start at fourd row
        c = 1 # column 'a'
        for row in rows:
            #print(row)
            for item in row:
                ws_bach.cell(row=r, column=c).value = item
                c += 1 # Column 'b'
            c = 1
            r += 1
        #monthly scrap report by day
        ws8=wb["scrap_type_machines"]
        Block.show_scrap_monthly_report_type_machines(self,year,month)
        #sql_query=Block.show_scrap_monthly_report_type_machines(self,year,month)
        #get_data = self.load_data(sql_query)        

        rows = cursor.fetchall()
        r = 3  # start at 33th row
        c = 1 # column 'a'
        for row in rows:
            #print(row)
        
            for item in row:
                ws8.cell(row=r, column=c).value = item
                c += 1 # Column 'b'
            c = 1
            r += 1
            #monthly machine report
        ws8=wb["scrap_days"]
        Block.show_scrap_monthly_report_by_days(self,year,month)
        #sql_query=Block.show_scrap_monthly_report_by_days(self,year,month)
        #get_data = self.load_data(sql_query)        
        rows = cursor.fetchall()
        r = 3  # start at 33th row
        c = 1 # column 'a'
        for row in rows:
            #print(row)
        
            for item in row:
                ws8.cell(row=r, column=c).value = item
                c += 1 # Column 'b'
            c = 1
            r += 1    
        #water content report
        ws1=wb["moisture_daily"]

        Block.show_water_content_daily(self,year,month,day,to_day)
        #sql_query=Block.show_water_content_daily(self,year,month,day,to_day)
        #get_data = self.load_data(sql_query)        
        get_data=cursor.fetchall()
        #get_data.set_index("serial", inplace=True) #put index
        
        #get_data=pd.DataFrame(get_data["id"])
        rows=get_data
        #rows = get_data[columns_quality]
        
        r = 4  # start at fourd row
        c = 1 # column 'a'
        for row in rows:
            #print(row)
            for item in row:
                ws1.cell(row=r, column=c).value = item
                c += 1 # Column 'b'
            c = 1
            r += 1
        #monthly output
        if monthly:
            ws2=wb["output"]
            
            sql_query=Block.show_monthly_report_ar(self,year,month,day,to_day)
            get_data = self.load_data(sql_query)
            rows = cursor.fetchall()
            r = 3  # start at third row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws2.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            #______yearly report
            #monthly mold report
            
            #filter on non conformity weights
                #part one low weight
            ws5=wb["wieght_report"]
            Block.monthly_report_ncr_weight_low(self,year,month)
            rows = cursor.fetchall()
            r = 12  # start at 12 row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws5.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
                #part tow hight weight
                ws5=wb["wieght_report"]
            Block.monthly_report_ncr_weight_hight(self,year,month)
            rows = cursor.fetchall()
            r = 33  # start at 33th row
            c = 1 # column 'a'
            for row in rows:
                
                for item in row:
                    ws5.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            #filter on non conformity ct
            ws6=wb["ct_report"]
            Block.monthly_report_ncr_ct(self,year,month)
            rows = cursor.fetchall()
            r = 11  # start at 11th row
            c = 1 # column 'a'
            for row in rows:
                
                for item in row:
                    ws6.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            #filter on on conformity scrap
            ws7=wb["scrap_report"]
            Block.monthly_report_ncr_scrap(self,year,month)
            rows= cursor.fetchall()
            r = 15  # start at 15th row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws7.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            #monthly machine report
            ws8=wb["scrap_machine"]
            Block.show_machine_monthly_report(self,year,month)
            rows = cursor.fetchall()
            r = 3  # start at 33th row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws8.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            
            
            #monthly machine report
            ws8=wb["scrap_machine_yearly"]
            Block.show_machine_yearly_report(self,year,month)
            rows = cursor.fetchall()
            r = 3  # start at 33th row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws8.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            
        
            ws2=wb["output_monthly"]
            Block.show_yearly_report_itemsByMonths(self,year,args)
            
            rows = cursor.fetchall()
            r = 3  # start at third row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws2.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            
            ws3=wb["year"]
            Block.show_machine_report_yearly(self,year,month)
            rows = cursor.fetchall()
            r = 3  # start at third row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws3.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            #scrap monthly
            ws3=wb["month"]
            Block.show_monthly_report_view_month(self,year,month)
            rows = cursor.fetchall()
            r = 3  # start at third row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws3.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            #yearly item report
            ws2=wb["output_yearly"]
            Block.items_report_arabic_custom_item(self,year,month,args)
            
            rows = cursor.fetchall()
            r = 3  # start at third row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws2.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            
            
            #monthly mold report by molds not items
            ws_output_molds=wb["output_molds"]
            Block.show_mnthly_report_molds(self,year,month)

            rows = cursor.fetchall()
            r = 3  # start at third row
            c = 1 # column 'a'
            for row in rows:
                #print(row)

                for item in row:
                    ws_output_molds.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1

            ws_output_molds_yearly=wb["output_mold_monthly"]
            Block.yearly_report_molds_byMonthes(self,year,args)

            rows = cursor.fetchall()
            r = 3  # start at third row
            c = 1 # column 'a'
            for row in rows:
                #print(row)

                for item in row:
                    ws_output_molds_yearly.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1

                #report for moldsy monthly
            
            ws_wieght_yearly=wb["output_molds_yearly"]
            Block.show_yearly_report_molds(self,year,month)
            rows = cursor.fetchall()
            r = 3  # start at 3th row
            c = 1 # column 'a'
            for row in rows:
                for item in row:
                    ws_wieght_yearly.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            
                #report of hight weight
            ws_wieght_yearly=wb["wieght_yearly"]
            Block.yearly_report_ncr_weight(self,year,month)
            rows = cursor.fetchall()
            r = 10  # start at 10th row
            c = 1 # column 'a'
            for row in rows:
                
                for item in row:
                    ws_wieght_yearly.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            #filter on non conformity ct
            ws_ct_yearly=wb["ct_yearly"]
            Block.yearly_report_ncr_ct(self,year,month)
            rows = cursor.fetchall()
            r = 11  # start at 11th row
            c = 1 # column 'a'
            for row in rows:
                
                for item in row:
                    ws_ct_yearly.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            #filter on on conformity scrap
            ws_scrap_yearly=wb["scrap_yearly"]
            Block.yearly_report_ncr_scrap(self,year,month)
            rows= cursor.fetchall()
            r = 15  # start at 15th row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws_scrap_yearly.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            #for material
            ws1=wb["materials"]
            Material.materialToPorduct(self,year)
            get_data=cursor.fetchall()
            #get_data.set_index("serial", inplace=True) #put index
            
            #get_data=pd.DataFrame(get_data["id"])
            rows=get_data
            #rows = get_data[columns_quality]
            
            r = 4  # start at fourd row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
                for item in row:
                    ws1.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1

            

        wb.save(writerFile)
    def molds_capabilty_study(self,mold,LSL,USL,eye_numbers,code,name,date):
        os.chdir(self.folder)
        
        
        Block.show_items(self)
        get_data=cursor.fetchall()
        #get_data.set_index("serial", inplace=True) #put index
        get_data=pd.read_excel("QC_daily_v2 - Copy.xlsx","items_spec")
        #get_data=pd.DataFrame(get_data["id"])
        data2=get_data[get_data['mold_id']==mold]
        data=data2
        #data=data2.split(",")
        director=data["mold_id"]
        print("kindly weignt , the data is generating")
        
        wb = xl.load_workbook(self.readfile1)
        ws2= wb.get_sheet_by_name('eye')
        

        #rows = get_data[columns_quality]
#        for items in len(data):

        #part=data[items]
        #item=[data["part_id"]==part]
        
        #ws2['h2']=self.sheet1
        
        ws2['a2']=code
        ws2['b2']=date
        ws2['k2']=name
        ws2['i9']=USL
        ws2['i10']=LSL
        for n in range(eye_numbers):
            ws=wb.copy_worksheet(ws2)   #copy new sheet
            r = 6  # row number
            c = 2 # column 'a'
            values=[]
            for i in range(50):
                
                random_value = randint(LSL,USL)            
                values.append(random_value)
                
                for item in values:
                    item=str(random_value)
                    ws.cell(row=r, column=c).value = item
                c = 2
                r += 1
            ws.title=self.writesheet
        wb.save(self.writefile)
    
    def export_report_daily_yearly(self,year,month,day,to_day,*args):
        os.chdir(self.folder)
        #new sheet
        
        wb = xl.load_workbook(self.readfile1)
        ws2= wb.get_sheet_by_name('input_yearly')
        
        ws=wb.copy_worksheet(ws2)   #copy new sheet

        #splite day
        
        
        ws.title="input"  #rename new sheet by the name
        
        
        sql_query=Block.get_daily_dataentry_items_yearly(self,year,month,day,to_day)
        get_data = self.load_data(sql_query)        

        list_item_size=get_data.shape[0]
        
        ws1=wb["input"]
        #create  the index sheet:
        r = 4  # start at 4th row
        c = 1 # column 'a'
        for row in range(0,list_item_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
            rows = get_data.iloc[row]
            for item in rows:
                ws1.cell(row=r, column=c).value = item
                c += 1 # Column 'd'
            c = 1
            r += 1   
            
        
        '''
        #for week
        ws1=wb["input-week"]
        
        Block.yearly_report_molds_byWeeks(self,year,month,day,to_day)
        get_data=cursor.fetchall()
        #get_data.set_index("serial", inplace=True) #put index
        
        #get_data=pd.DataFrame(get_data["id"])
        rows=get_data
        #rows = get_data[columns_quality]
        
        r = 4  # start at fourd row
        c = 1 # column 'a'
        for row in rows:
            #print(row)
            for item in row:
                ws1.cell(row=r, column=c).value = item
                c += 1 # Column 'b'
            c = 1
            r += 1

        '''
        wb.save(year+"-QC_molds_daily_yearly_v3.xlsx")
        #to save excel sheet by columns original columns
        writer_report = pd.ExcelWriter("QC_molds_daily_yearly_v3.xlsx")
        writer_report.save()

    def daily_molds(self,year,month,day):
        '''
            1- to get summary report for each day
        '''        

        os.chdir(self.folder)

        wb = xl.load_workbook(self.readfile1)
        sql_query=Block.get_daily_dataentry_items(self,year,month,day)
        mold_analysis4 = self.load_data(sql_query)        

        last_year=int(mold_analysis4["year"].max())
        print("_________daily report___________for year________",last_year,type(last_year))       
        mold_analysis_bool4=mold_analysis4["year"]==last_year
        mold_analysis3=mold_analysis4[mold_analysis_bool4]
        
        last_month=int(mold_analysis3["month"].max())
        mold_analysis_bool3=mold_analysis3["month"]==last_month
        mold_analysis2=mold_analysis3[mold_analysis_bool3]
        
        #dateDay3=mold_analysis2['date_day'].tail(1)
        print("_________daily report___________for month________",last_month,type(last_month))       
        mold_days=mold_analysis2["day"].count()

        import datetime

        d = datetime.datetime(int(year), int(month), int(day))

        date_day=d.date()
      

        daily_analysis1_bool=mold_analysis2["day"]==date_day

        daily_analysis1=mold_analysis2[daily_analysis1_bool]
        #validate input 

        #for fix nan error in ct
        daily_analysis = daily_analysis1.dropna(subset=['c_t_actually'])#remove all numric data
        daily_analysis = daily_analysis.dropna(subset=['c_t_actually'])#drop And for remove all rows with NaNs in column x use dropna: 
        daily_analysis["c_t_actually"]=daily_analysis["c_t_actually"]#Last convert values to ints:

        report_forCt=daily_analysis.groupby(["machine_id","mold_name"])["c_t_standard_per_second","standard_rate_hour","rat_actually","c_t_actually"].mean()
        
        
        print("___________daily_analysis   standard_dry_weight_from",mold_analysis2[["standard_dry_weight_from"]])

        print("___________report_forCt   data",report_forCt)
        #report=pd.DataFrame()
        
        
        mold_count=report_forCt["rat_actually"].count()
        #cycle timpe report
        
        c_t=report_forCt[["c_t_standard_per_second","standard_rate_hour","rat_actually","c_t_actually"]]
        
        c_t_nonconfomity=c_t[c_t['rat_actually']*1.05<c_t['standard_rate_hour']]
        c_t_nonconfomity_count=c_t_nonconfomity["rat_actually"].count()
        
        
        if c_t_nonconfomity_count==0:  # for ignor impty index error
            c_t_nonconfomity=pd.DataFrame(index=[0])
            c_t_nonconfomity["mold_name"]=0
            c_t_nonconfomity["c_t_standard_per_second"]=0
            c_t_nonconfomity["standard_rate_hour"]=0
            c_t_nonconfomity["rat_actually"]=0
            
        ct_ok=mold_count-c_t_nonconfomity_count
        c_t_nonconfomity["c_t_nonconfomity_count"]=c_t_nonconfomity_count
        c_t_nonconfomity["ct_ok"]=ct_ok
        #weight report
       #fix whistespaces in column names
        weight_cl=["machine_id","mold_name","standard_dry_weight_from","standard_dry_weight_to","average_dry_weight"]
        wieght3=daily_analysis1[weight_cl]
        #new_data = pd.DataFrame()
        #new_data['tsneY'] = df['tsneY'].values.tolist()
        print('weight3',wieght3)
        wieght2=daily_analysis.groupby(["machine_id","mold_name"])["standard_dry_weight_from","standard_dry_weight_to","average_dry_weight"].mean()
        print ("test_____________wieght2[standard_dry_weight_from",wieght2,wieght2.info())

        #filter low weithrs

        weight_nonconfomity_low=wieght2[wieght2['average_dry_weight'] <= wieght2["standard_dry_weight_from"]]

        weight_nonconfomity_high=wieght2[wieght2['average_dry_weight']>wieght2["standard_dry_weight_to"]]

        wieght=weight_nonconfomity_high
        wieght=wieght.append(weight_nonconfomity_low)

        weight_nonconfomity=weight_nonconfomity_high
        weight_nonconfomity_count=weight_nonconfomity["average_dry_weight"].count()
        if weight_nonconfomity_count==0:  # for ignor impty index error        
            weight_nonconfomity=pd.DataFrame(index=[0])
        if weight_nonconfomity_count>1:  # for ignor impty index error        
            weight_nonconfomity=pd.DataFrame()##error we muast select index of groupby
        
        weight_nonconfomity=weight_nonconfomity.append(weight_nonconfomity_low)    #for append the light weights
        
        weight_nonconfomity['weight_nonconfomity_count']=weight_nonconfomity_count
        weight_ok=mold_count-weight_nonconfomity_count
        weight_nonconfomity['weight_ok']=weight_ok
        
        ##screap report by parts##_________
                #scap by parts
        #report=pd.DataFrame()
        scrap5=daily_analysis1[["machine_type","machine_id",'number_scrab_by_item','gross_production','scrap_percent_by_item',"mold_name","scrap_weight_kg","production_weight_kg","product_name","scrabe_standard","average_dry_weight"]]
        
        #scrap5=report.append(scrap5)#we need fix rong calucate sacrap percent for 
        
        scrap_newmachines=scrap5[scrap5["machine_type"]=="new_machine"]
        scrap_oldmachines=scrap5[scrap5["machine_type"]=="old_machine"]
        
        scrap_part_new=scrap_newmachines["number_scrab_by_item"].sum()
        production_part_new=scrap_newmachines["gross_production"].sum()
        
        scrap_percent_new=scrap_part_new/production_part_new * 100
        scrap_part=scrap5["number_scrab_by_item"].sum()
        production_set=scrap5["gross_production"].sum()
        scrap_percent=(scrap_part/production_set * 100).astype(int)
        scrap4=scrap5[scrap5["average_dry_weight"]>1]      #for filter the using machines only
        
    
        scrap3=scrap4[scrap4["scrap_percent_by_item"]>scrap4["scrabe_standard"]]
        
        scrap2=scrap3[scrap3["gross_production"]>scrap3["number_scrab_by_item"]]# for filter any machines didn't create production
        
        scrap=scrap2
        print("_______________scrap____________________",scrap)
        scrap["scrab_parts_new_machine"]=scrap_part_new
        scrap["production_parts_new_machine"]=production_part_new
        scrap["production_parts_all"]=production_set
        scrap["scrap_percent_new_machine"]=scrap_percent_new
        scrap["scrap_percent_all"]=scrap_percent
        scrap["scrab_parts_all"]=scrap_part
        scrap_nonconfomity_count=scrap["number_scrab_by_item"].count()        
        ##scrap for molds___________________
        '''must be in excel sheet (input) not showing null value for not mistacks in average result'''
        scrap_molds=scrap5.groupby(["machine_type","machine_id","mold_name","scrabe_standard"])['number_scrab_by_item'].sum()
        scrap_molds["gross_production"]=scrap5.groupby(["machine_type","machine_id","mold_name","scrabe_standard"])['gross_production'].sum()
        scrap_molds["scrap_weight_kg"]=scrap5.groupby(["machine_type","machine_id","mold_name","scrabe_standard"])['scrap_weight_kg'].sum()
        scrap_molds["production_weight_kg"]=scrap5.groupby(["machine_type","machine_id","mold_name","scrabe_standard"])['production_weight_kg'].sum()
        scrap_molds["number_scrab_by_item"]=scrap5.groupby(["machine_type","machine_id","mold_name","scrabe_standard"])['number_scrab_by_item'].sum()
        production_set_molds=scrap_molds["gross_production"].sum()
        scrap_set_molds=scrap_molds["number_scrab_by_item"].sum()
        scrap_percent_molds=(scrap_set_molds.astype(int)/production_set_molds.astype(int))*100
        scrap_molds["percent"]=(scrap_molds["number_scrab_by_item"]/scrap_molds["gross_production"])*100
        ##scrap for items(as item master)___
        scrap_items=scrap5.groupby(["machine_type","machine_id","product_name"])['number_scrab_by_item'].sum()
        
        scrap_items["gross_production"]=scrap5.groupby(["machine_type","machine_id","product_name"])['gross_production'].sum()
        #scrap_items["number_day_use"]=scrap5.groupby(["machine_type","machine_id","product_name"])['number_day_use'].mean()
        scrap_items["scrap_weight_kg"]=scrap5.groupby(["machine_type","machine_id","product_name"])['scrap_weight_kg'].sum()
        scrap_items["production_weight_kg"]=scrap5.groupby(["machine_type","machine_id","product_name"])['production_weight_kg'].sum()
        scrap_items["number_scrab_by_item"]=scrap5.groupby(["machine_type","machine_id","product_name"])['number_scrab_by_item'].sum()
        scrap_items["percent"]=((scrap_items["number_scrab_by_item"])/(scrap_items["gross_production"]))*100

        scrap_percent_new=scrap_part_new/production_part_new * 100
        ##scrap non conformity_______________________
        
        
        if scrap_nonconfomity_count>0:  # for ignor impty index error
        #    scrap=pd.DataFrame(index=[0])
            scrap_nonconfomity=scrap2.groupby(["machine_id","mold_name"])['number_scrab_by_item'].sum()
            #for ignor error when srcap is 0
            
            
        else:
            #if scrap_nonconfomity.datatypt==0:  # for ignor impty index error
            scrap_nonconfomity=pd.DataFrame(index=[0])
            scrap_nonconfomity["mold_name"]=0
        print("_____________test__________scrap nonconfromity____",scrap_nonconfomity)
        #extract excel report
        writer_report = pd.ExcelWriter("QC_molds_daily_archive.xlsx")
        mold_analysis2.to_excel(writer_report,"input", index=False)
        writer_report.save()

        writer = pd.ExcelWriter("day_analysis2.xlsx")

        weight_nonconfomity.to_excel(writer,"weight_ncr")
        
        c_t_nonconfomity.to_excel(writer,"c.t")
        scrap.to_excel(writer,"scrap")
        scrap_molds.to_excel(writer,"scrap_molds",merge_cells=False)
        scrap_nonconfomity.to_excel(writer,"scrap_ncr")
        scrap_items.to_excel(writer,"scrap_items",merge_cells=False)
    
#        wieght2.to_excel(writer,merge_cells=False)
        daily_analysis1.to_excel(writer,"input_molds", index=False)
        
        writer.save()
        
        #extract daily excel report by formating

        wb = load_workbook("QC_molds_daily_summary.xlsx")
        ws2= wb.get_sheet_by_name('daily')
        
        ws=wb.copy_worksheet(ws2)   #copy new sheet

        #splite day
        
        
        ws.title=str(day)  #rename new sheet by the name
        
                
        # Data can be assigned directly to cells
        ws['b1'] = day            #day for molds report
        ws['c1'] = last_month    #month for molds report
        ws['d1'] = last_year    #year for molds report
        ws['A4'] = scrap_percent           #scrap on all machie
        ws['B4'] = scrap_percent_new    #scrap on new macine
        print ("c_t_nonconfomity for sheet" , c_t_nonconfomity)
        ws['a14'] =c_t_nonconfomity.iloc[0][5]                #ct pass 
        ws['b14'] =c_t_nonconfomity.iloc[0][4]                #ct not acceptable
        
        ws['a25'] = weight_nonconfomity.iloc[0][4]             #weight pass 
        ws['b25'] =weight_nonconfomity.iloc[0][3]
        #if weight_nonconfomity_high>=1:  # for ignor impty index error        
        print ("weight_nonconfomity_low for sheet" , weight_nonconfomity)
        #ws['b25'] =weight_nonconfomity_low.iloc[0][2]            #weights low not acceptable
        #else:
        #    ws['b25'] =0
        ws['a35'] = weight_nonconfomity.iloc[0][4]             #weight pass 
        #if weight_nonconfomity_high>=1:  # for ignor impty index error        
        #ws['b35'] =weight_nonconfomity_high.iloc[0][2]            #weights hig not acceptable
        #else:
        #    ws['b35'] =0           #weights hig not acceptable
        #_______
        
        #ws['a35'] = weight_nonconfomity.iloc[0][3]             #weight pass 
        #ws['b25'] =weight_nonconfomity.iloc[0][2]               #weight not pass

        #to select index of columns for scraps
        if scrap_nonconfomity_count>=1:
            rows = scrap_nonconfomity.index
            r = 4  # start at 10th row
            c = 3 # column 'c'
            for row in rows:       
                for item in row:
                    ws.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 3
                r += 1

        #to select index of columns to weights
        
        if c_t_nonconfomity_count>=1:  # for ignor impty index error        
            rows = c_t_nonconfomity.index
            #rows = weight_nonconfomity.index
            r = 14  # start at 10th row
            c = 3 # column 'c'
            
            for row in rows:       
                for item in row:
                    ws.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 3
                r += 1
        
        #to select index of columns to light weights 
        if weight_nonconfomity_count>=1:  # for ignor impty index error        
            rows = weight_nonconfomity_low.index
            #rows = weight_nonconfomity.index
            r = 25  # start at 10th row
            c = 3 # column 'c'
            
            for row in rows:       
                for item in row:
                    ws.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 3
                r += 1
        #to select index of columns to hight weights 
        if weight_nonconfomity_count>=1:  # for ignor impty index error        
            rows = weight_nonconfomity_high.index
            #rows = weight_nonconfomity.index
            r = 35  # start at 10th row
            c = 3 # column 'c'
            
            for row in rows:       
                for item in row:
                    ws.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 3
                r += 1
        
        wb.save("QC_molds_daily_summary.xlsx")
        #__________________archinve work boook monthly and daily report________________________________________        
        #__________________archinve work boook monthly and daily report general________________________________________
        #input report
        wb_formats=load_workbook("format_QC_reports_v2.xlsx")    
        ws_input= wb_formats.get_sheet_by_name('input')
        #ws_input=wb_formats.copy_worksheet(ws_input2)   #copy new sheet
        ws_input.sheet_view.zoomScale = 60 #zoom set

        column_size_input=pd.DataFrame(mold_analysis2,columns=["month"]).shape[0]
        #print("______test___")
        #print(mold_analysis2)
        
        r = 4  # start at 4th row
        c = 1 # column 'a'
        
        for row in range(0,column_size_input):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
            rows = mold_analysis2.iloc[row]
            for item in rows:
                ws_input.cell(row=r, column=c).value = item
                c += 1 # Column 'd'
            c = 1
            r += 1
        #__________________archinve work boook monthly and daily report for specific mold
        
        
        wb_formats.save("QC_molds_daily_archive.xlsx")    
        
    def monthly_molds(self,writerFile,year,month,day,to_day,*args,monthly=True,daily=True):
        '''
        mold report for 
            1- to get summary report for each day
            2- to get daily report 
            3- monthly repots ended by QC_molds_daily_archive_v3
        
        '''        

        os.chdir(self.folder)

        wb = xl.load_workbook(self.readfile1)
        
        sql_query=Block.get_daily_dataentry_items(self,year,month,day)

        get_data = self.load_data(sql_query)        
        

        last_year=int(get_data["year"].max())
        print("_________daily report___________for year________",last_year,type(last_year))       
        mold_analysis_bool4=get_data["year"]==last_year
        mold_analysis3=get_data[mold_analysis_bool4]
        
        last_month=int(mold_analysis3["month"].max())
        mold_analysis_bool3=mold_analysis3["month"]==last_month
        mold_analysis2=mold_analysis3[mold_analysis_bool3]
        
        #dateDay3=mold_analysis2['date_day'].tail(1)
        print("_________daily report___________for month________",last_month,type(last_month))       
        mold_days=mold_analysis2["day"].count()

        import datetime

        d = datetime.datetime(int(year), int(month), int(day))

        date_day=d.date()
      
        
        daily_analysis1_bool=mold_analysis2["day"]==date_day

        daily_analysis1=mold_analysis2[daily_analysis1_bool]
        #validate input 

        daily_analysis=mold_analysis2
        #daily_analysis=daily_analysis[column_monthly_report]       
        
        print("daily_analysis for day",day,daily_analysis)
        print("date",year,month,day)
        dry_weight3=daily_analysis[columns_weight]
        dry_weight_bool=dry_weight3['average_dry_weight'].notnull()
        dry_weight2=dry_weight3[dry_weight_bool]

        scrap4=daily_analysis[columns_machine]
        scrap_bool=scrap4["gross_production"].notnull()
        scrap3=scrap4[scrap_bool]
        scrap2=scrap3
        molds_rate3=daily_analysis[columns_cycle_time]
        molds_rate_bool=molds_rate3["c_t_actually"].notnull()
        molds_rate2=molds_rate3[molds_rate_bool]
        print("please weight1")
        #scrap
        scrap=scrap2.groupby(["product_name","product_code","scrabe_standard"])['number_scrab_by_item',
        "sum_scrabe_no_parts","gross_production","number_day_use"].sum()
        
        scrap["scrap_percent_by_item"]=scrap['number_scrab_by_item']/scrap['gross_production']


        scrap_product_machine=scrap2.groupby(["product_name","product_code","machine_id","scrabe_standard"])['number_scrab_by_item',
        "sum_scrabe_no_parts","gross_production","number_day_use"].sum()

        scrap_product_machine["scrap_percent_by_item"]=scrap_product_machine['number_scrab_by_item']/scrap_product_machine['gross_production']
        
        print("_____scrap____",scrap_product_machine)

        scrap_machine_product=scrap2.groupby(["machine_id","scrabe_standard","product_name","product_code"])['number_scrab_by_item',
        "sum_scrabe_no_parts","gross_production","number_day_use"].sum()
        scrap_machine_product["scrap_percent_by_item"]=scrap_machine_product['number_scrab_by_item']/scrap_machine_product['gross_production']
        
        machines=scrap2.groupby(["machine_id","scrabe_standard","machine_type"])['number_scrab_by_item',
        'sum_scrabe_no_parts',"gross_production","number_day_use"].sum()
        machines["scrap_percent_by_item"]=machines['number_scrab_by_item']/machines['gross_production']
        scrap_bool2=scrap2["scrap_percent_by_item"]>scrap2["scrabe_standard"]
        scrap_ncr=scrap2[scrap_bool2]
        print("please weight2")
        #production rate report
        
        molds_rate=molds_rate2.groupby(["product_name","product_code","set","standard_rate_hour"
        ,"c_t_standard_per_second"])["rat_actually","c_t_actually"].mean()
        #molds_rate_bools_ncr=molds_rate["rat_actually"]>molds_rate["standard_rate_hour"]
        #molds_rate_ncr=molds_rate[molds_rate_bools_ncr]
        dry_weight=dry_weight2.groupby(["product_name","product_code","standard_dry_weight",
        "standard_dry_weight_from","standard_dry_weight_to"])["average_dry_weight","average_wet_weight"].mean()
        

        # export
        output_average=daily_analysis.groupby(["product_name","product_code","standard_dry_weight",
        "standard_dry_weight_from","standard_dry_weight_to","standard_rate_hour"
        ,"c_t_standard_per_second","scrabe_standard"])["average_dry_weight","average_wet_weight","rat_actually","c_t_actually"].mean()
        
        output_aggrigate=daily_analysis.groupby(["product_name","product_code","standard_dry_weight",
        "standard_dry_weight_from","standard_dry_weight_to","standard_rate_hour"
        ,"c_t_standard_per_second","scrabe_standard"])['number_scrab_by_item',
        "sum_scrabe_no_parts","gross_production","number_day_use"].sum()

        output=output_average
        output.append(output_aggrigate)
        
        #validation data
        #scap validation
        product_parts_input=daily_analysis["gross_production"].sum
        
        scrab_set_input=daily_analysis["number_scrab_by_item"].sum
        #scrab_input=product_parts_input + product_set_input + scrab_parts_input + scrab_set_input

        product_parts_outbut=machines["gross_production"].sum
        
        
        scrab_set_outbut=machines["number_scrab_by_item"].sum
        #scrab_output=product_parts_outbut + product_set_outbut +  scrab_parts_outbut + scrab_set_outbut

     
        writer = pd.ExcelWriter("QC_molds_monthly_until.xlsx")
    
   
        daily_analysis.rename(columns={c:c.lower() for c in col_rename})      

        daily_analysis.to_excel(writer,'input', index=False)

        scrap.rename(columns={c:c.lower() for c in col_rename})
        scrap.to_excel(writer,'scrap_product', index=True)

        scrap_product_machine.rename(columns={c:c.lower() for c in col_rename})
        scrap_product_machine.to_excel(writer,'scrap_product_machines', index=True)

        scrap_machine_product.rename(columns={c:c.lower() for c in col_rename})
        scrap_machine_product.to_excel(writer,'scrap_machines_product', index=True)
        
        machines.rename(columns={c:c.lower() for c in col_rename})
        machines.to_excel(writer,"scrap_machines", index=True)
        
        #dry_weight.rename(columns={c:c.lower() for c in col_rename})
        dry_weight.to_excel(writer,'weights', index=True)
        molds_rate.rename(columns={c:c.lower() for c in col_rename})
        molds_rate.to_excel(writer,'c_t', index=True)
        output.to_excel(writer,"output",index=True)
        writer.save()

        print("for the days")
        print(daily_analysis["day"].unique())
    #   ncrs
    #for fix nan error in ct
        daily_analysis = daily_analysis1.dropna(subset=['c_t_actually'])#remove all numric data
        daily_analysis = daily_analysis.dropna(subset=['c_t_actually'])#drop And for remove all rows with NaNs in column x use dropna: 
        daily_analysis["c_t_actually"]=daily_analysis["c_t_actually"]#Last convert values to ints:

        report_forCt=daily_analysis.groupby(["machine_id","mold_name"])["c_t_standard_per_second","standard_rate_hour","rat_actually","c_t_actually"].mean()
        
        
        print("___________daily_analysis   standard_dry_weight_from",mold_analysis2[["standard_dry_weight_from"]])

        print("___________report_forCt   data",report_forCt)
        #report=pd.DataFrame()
        
        
        mold_count=report_forCt["rat_actually"].count()
        #cycle timpe report
        
        c_t=report_forCt[["c_t_standard_per_second","standard_rate_hour","rat_actually","c_t_actually"]]
        
        c_t_nonconfomity=c_t[c_t['rat_actually']*1.05<c_t['standard_rate_hour']]
        c_t_nonconfomity_count=c_t_nonconfomity["rat_actually"].count()
        
        
        if c_t_nonconfomity_count==0:  # for ignor impty index error
            c_t_nonconfomity=pd.DataFrame(index=[0])
            c_t_nonconfomity["mold_name"]=0
            c_t_nonconfomity["c_t_standard_per_second"]=0
            c_t_nonconfomity["standard_rate_hour"]=0
            c_t_nonconfomity["rat_actually"]=0
            
        ct_ok=mold_count-c_t_nonconfomity_count
        c_t_nonconfomity["c_t_nonconfomity_count"]=c_t_nonconfomity_count
        c_t_nonconfomity["ct_ok"]=ct_ok
        #weight report
       #fix whistespaces in column names
        weight_cl=["machine_id","mold_name","standard_dry_weight_from","standard_dry_weight_to","average_dry_weight"]
        wieght3=daily_analysis1[weight_cl]
        #new_data = pd.DataFrame()
        #new_data['tsneY'] = df['tsneY'].values.tolist()
        print('weight3',wieght3)
        wieght2=daily_analysis.groupby(["machine_id","mold_name"])["standard_dry_weight_from","standard_dry_weight_to","average_dry_weight"].mean()
        print ("test_____________wieght2[standard_dry_weight_from",wieght2,wieght2.info())

        #filter low weithrs

        weight_nonconfomity_low=wieght2[wieght2['average_dry_weight'] <= wieght2["standard_dry_weight_from"]]

        weight_nonconfomity_high=wieght2[wieght2['average_dry_weight']>wieght2["standard_dry_weight_to"]]

        wieght=weight_nonconfomity_high
        wieght=wieght.append(weight_nonconfomity_low)

        weight_nonconfomity=weight_nonconfomity_high
        weight_nonconfomity_count=weight_nonconfomity["average_dry_weight"].count()
        if weight_nonconfomity_count==0:  # for ignor impty index error        
            weight_nonconfomity=pd.DataFrame(index=[0])
        if weight_nonconfomity_count>1:  # for ignor impty index error        
            weight_nonconfomity=pd.DataFrame()##error we muast select index of groupby
        
        weight_nonconfomity=weight_nonconfomity.append(weight_nonconfomity_low)    #for append the light weights
        
        weight_nonconfomity['weight_nonconfomity_count']=weight_nonconfomity_count
        weight_ok=mold_count-weight_nonconfomity_count
        weight_nonconfomity['weight_ok']=weight_ok
        
        ##screap report by parts##_________
                #scap by parts
        #report=pd.DataFrame()
        scrap5=daily_analysis1[["machine_type","machine_id",'number_scrab_by_item','gross_production','scrap_percent_by_item',"mold_name","scrap_weight_kg","production_weight_kg","product_name","scrabe_standard","average_dry_weight"]]
        
        #scrap5=report.append(scrap5)#we need fix rong calucate sacrap percent for 
        
        scrap_newmachines=scrap5[scrap5["machine_type"]=="new_machine"]
        scrap_oldmachines=scrap5[scrap5["machine_type"]=="old_machine"]
        
        scrap_part_new=scrap_newmachines["number_scrab_by_item"].sum()
        production_part_new=scrap_newmachines["gross_production"].sum()
        
        scrap_percent_new=scrap_part_new/production_part_new * 100
        scrap_part=scrap5["number_scrab_by_item"].sum()
        production_set=scrap5["gross_production"].sum()
        scrap_percent=scrap_part/production_set * 100
        scrap4=scrap5[scrap5["average_dry_weight"]>1]      #for filter the using machines only
        
    
        scrap3=scrap4[scrap4["scrap_percent_by_item"]>scrap4["scrabe_standard"]]
        
        scrap2=scrap3[scrap3["gross_production"]>scrap3["number_scrab_by_item"]]# for filter any machines didn't create production
        
        scrap=scrap2
        print("_______________scrap____________________",scrap)
        scrap["scrab_parts_new_machine"]=scrap_part_new
        scrap["production_parts_new_machine"]=production_part_new
        scrap["production_parts_all"]=production_set
        scrap["scrap_percent_new_machine"]=scrap_percent_new
        scrap["scrap_percent_all"]=scrap_percent
        scrap["scrab_parts_all"]=scrap_part
        scrap_nonconfomity_count=scrap["number_scrab_by_item"].count()        
        ##scrap for molds___________________
        '''must be in excel sheet (input) not showing null value for not mistacks in average result'''
        scrap_molds=scrap5.groupby(["machine_type","machine_id","mold_name","scrabe_standard"])['number_scrab_by_item'].sum()
        scrap_molds["gross_production"]=scrap5.groupby(["machine_type","machine_id","mold_name","scrabe_standard"])['gross_production'].sum()
        scrap_molds["scrap_weight_kg"]=scrap5.groupby(["machine_type","machine_id","mold_name","scrabe_standard"])['scrap_weight_kg'].sum()
        scrap_molds["production_weight_kg"]=scrap5.groupby(["machine_type","machine_id","mold_name","scrabe_standard"])['production_weight_kg'].sum()
        scrap_molds["number_scrab_by_item"]=scrap5.groupby(["machine_type","machine_id","mold_name","scrabe_standard"])['number_scrab_by_item'].sum()
        production_set_molds=scrap_molds["gross_production"].sum()
        scrap_set_molds=scrap_molds["number_scrab_by_item"].sum()
        scrap_percent_molds=(scrap_set_molds.astype(int)/production_set_molds.astype(int))*100
        scrap_molds["percent"]=(scrap_molds["number_scrab_by_item"]/scrap_molds["gross_production"])*100
        ##scrap for items(as item master)___
        scrap_items=scrap5.groupby(["machine_type","machine_id","product_name"])['number_scrab_by_item'].sum()
        
        scrap_items["gross_production"]=scrap5.groupby(["machine_type","machine_id","product_name"])['gross_production'].sum()
        #scrap_items["number_day_use"]=scrap5.groupby(["machine_type","machine_id","product_name"])['number_day_use'].mean()
        scrap_items["scrap_weight_kg"]=scrap5.groupby(["machine_type","machine_id","product_name"])['scrap_weight_kg'].sum()
        scrap_items["production_weight_kg"]=scrap5.groupby(["machine_type","machine_id","product_name"])['production_weight_kg'].sum()
        scrap_items["number_scrab_by_item"]=scrap5.groupby(["machine_type","machine_id","product_name"])['number_scrab_by_item'].sum()
        scrap_items["percent"]=((scrap_items["number_scrab_by_item"])/(scrap_items["gross_production"]))*100

        scrap_percent_new=scrap_part_new/production_part_new * 100
        ##scrap non conformity_______________________
        
        
        if scrap_nonconfomity_count>0:  # for ignor impty index error
        #    scrap=pd.DataFrame(index=[0])
            scrap_nonconfomity=scrap2.groupby(["machine_id","mold_name"])['number_scrab_by_item'].sum()
            #for ignor error when srcap is 0
            
            
        else:
            #if scrap_nonconfomity.datatypt==0:  # for ignor impty index error
            scrap_nonconfomity=pd.DataFrame(index=[0])
            scrap_nonconfomity["mold_name"]=0
        print("_____________test__________scrap nonconfromity____",scrap_nonconfomity)
        #extract excel report
        writer_report = pd.ExcelWriter("QC_molds_daily_archive.xlsx")
        mold_analysis2.to_excel(writer_report,"input", index=False)
        writer_report.save()

        writer = pd.ExcelWriter("day_analysis2.xlsx")

        weight_nonconfomity.to_excel(writer,"weight_ncr")
        
        c_t_nonconfomity.to_excel(writer,"c.t")
        scrap.to_excel(writer,"scrap")
        scrap_molds.to_excel(writer,"scrap_molds",merge_cells=False)
        scrap_nonconfomity.to_excel(writer,"scrap_ncr")
        scrap_items.to_excel(writer,"scrap_items",merge_cells=False)
    
#        wieght2.to_excel(writer,merge_cells=False)
        daily_analysis1.to_excel(writer,"input_molds", index=False)
        
        writer.save()
        
        #extract daily excel report by formating
        if daily:
            wb_summary = load_workbook("QC_molds_daily_summary.xlsx")
            ws2= wb_summary.get_sheet_by_name('daily')
            
            ws=wb_summary.copy_worksheet(ws2)   #copy new sheet

            #splite day
            #lastDay = [daily_analysis1["day_date"].split('-')[2] for d in daily_analysis1.Date]
            #lastDay=pd.to_DateFrame(daily_analysis1,columns=["day_date"])
            #daily_analysis1.lastDay=pd.to_datetime(lastDay.Datetime,format='%d')
            
            ws.title=str(day)  #rename new sheet by the name
            
                    
            # Data can be assigned directly to cells
            ws['b1'] = day            #day for molds report
            ws['c1'] = last_month    #month for molds report
            ws['d1'] = last_year    #year for molds report
            ws['A4'] = scrap_percent           #scrap on all machie
            ws['B4'] = scrap_percent_new    #scrap on new macine
            print ("c_t_nonconfomity for sheet" , c_t_nonconfomity)
            ws['a14'] =c_t_nonconfomity.iloc[0][5]                #ct pass 
            ws['b14'] =c_t_nonconfomity.iloc[0][4]                #ct not acceptable
            
            ws['a25'] = weight_nonconfomity.iloc[0][4]             #weight pass 
            ws['b25'] =weight_nonconfomity.iloc[0][3]
            #if weight_nonconfomity_high>=1:  # for ignor impty index error        
            print ("weight_nonconfomity_low for sheet" , weight_nonconfomity)
            #ws['b25'] =weight_nonconfomity_low.iloc[0][2]            #weights low not acceptable
            #else:
            #    ws['b25'] =0
            ws['a35'] = weight_nonconfomity.iloc[0][4]             #weight pass 
            #if weight_nonconfomity_high>=1:  # for ignor impty index error        
            #ws['b35'] =weight_nonconfomity_high.iloc[0][2]            #weights hig not acceptable
            #else:
            #    ws['b35'] =0           #weights hig not acceptable
            #_______
            
            #ws['a35'] = weight_nonconfomity.iloc[0][3]             #weight pass 
            #ws['b25'] =weight_nonconfomity.iloc[0][2]               #weight not pass

            #to select index of columns for scraps
            if scrap_nonconfomity_count>=1:
                rows = scrap_nonconfomity.index
                r = 4  # start at 10th row
                c = 3 # column 'c'
                for row in rows:       
                    for item in row:
                        ws.cell(row=r, column=c).value = item
                        c += 1 # Column 'd'
                    c = 3
                    r += 1

            #to select index of columns to weights
            
            if c_t_nonconfomity_count>=1:  # for ignor impty index error        
                rows = c_t_nonconfomity.index
                #rows = weight_nonconfomity.index
                r = 14  # start at 10th row
                c = 3 # column 'c'
                
                for row in rows:       
                    for item in row:
                        ws.cell(row=r, column=c).value = item
                        c += 1 # Column 'd'
                    c = 3
                    r += 1
            
            #to select index of columns to light weights 
            if weight_nonconfomity_count>=1:  # for ignor impty index error        
                rows = weight_nonconfomity_low.index
                #rows = weight_nonconfomity.index
                r = 25  # start at 10th row
                c = 3 # column 'c'
                
                for row in rows:       
                    for item in row:
                        ws.cell(row=r, column=c).value = item
                        c += 1 # Column 'd'
                    c = 3
                    r += 1
            #to select index of columns to hight weights 
            if weight_nonconfomity_count>=1:  # for ignor impty index error        
                rows = weight_nonconfomity_high.index
                #rows = weight_nonconfomity.index
                r = 35  # start at 10th row
                c = 3 # column 'c'
                
                for row in rows:       
                    for item in row:
                        ws.cell(row=r, column=c).value = item
                        c += 1 # Column 'd'
                    c = 3
                    r += 1
            
            wb_summary.save("QC_molds_daily_summary.xlsx")
        
    #__________________________________merge
        sheets_daily=["input_daily"]
        sheets_monthly=["input_daily","output","wieght_report","ct_report","scrap_report"]

        #for copy sheet
        #ws2= wb.get_sheet_by_name('input_daily')
        #ws1=wb.copy_worksheet(ws2)   #copy new sheet
        #splite day
        #ws1.title="input"  #rename new sheet by the name
        
        list_item_size=get_data.shape[0]
        ws1=wb["input_daily"]
        #create  the index sheet:
        r = 4  # start at 4th row
        c = 1 # column 'a'
        for row in range(0,list_item_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
            rows = get_data.iloc[row]
            for item in rows:
                ws1.cell(row=r, column=c).value = item
                c += 1 # Column 'd'
            c = 1
            r += 1   
        
        #material by silo
        ws_bach=wb["material_daily"]
        Material.material_bySilo_daily(self,year,month,day,to_day)
        get_data=cursor.fetchall()
        rows=get_data      
        r = 4  # start at fourd row
        c = 1 # column 'a'
        for row in rows:
            #print(row)
            for item in row:
                ws_bach.cell(row=r, column=c).value = item
                c += 1 # Column 'b'
            c = 1
            r += 1
        #Bache input
        ws_bach=wb["batches"]
        Block.show_monthly_Baches(self,year,month)
        
#        sql_query=Block.show_monthly_Baches(self,year,month)
#       get_data = self.load_data(sql_query)        

        get_data=cursor.fetchall()
        rows=get_data
        #rows = get_data[columns_quality]
        
        r = 4  # start at fourd row
        c = 1 # column 'a'
        for row in rows:
            #print(row)
            for item in row:
                ws_bach.cell(row=r, column=c).value = item
                c += 1 # Column 'b'
            c = 1
            r += 1
        #monthly scrap report by day
        ws8=wb["scrap_type_machines"]

        rows = scrap_machine_product
        r = 3  # start at 33th row
        c = 1 # column 'a'
        for row in rows:
            #print(row)
        
            for item in row:
                ws8.cell(row=r, column=c).value = item
                c += 1 # Column 'b'
            c = 1
            r += 1
            #monthly machine report
        ws8=wb["scrap_days"]
        Block.show_scrap_monthly_report_by_days(self,year,month)
        #sql_query=Block.show_scrap_monthly_report_by_days(self,year,month)
        #get_data = self.load_data(sql_query)        
        rows = cursor.fetchall()
        r = 3  # start at 33th row
        c = 1 # column 'a'
        for row in rows:
            #print(row)
        
            for item in row:
                ws8.cell(row=r, column=c).value = item
                c += 1 # Column 'b'
            c = 1
            r += 1    
        #water content report
        ws1=wb["moisture_daily"]

        Block.show_water_content_daily(self,year,month,day,to_day)
        #sql_query=Block.show_water_content_daily(self,year,month,day,to_day)
        #get_data = self.load_data(sql_query)        
        get_data=cursor.fetchall()
        #get_data.set_index("serial", inplace=True) #put index
        
        #get_data=pd.DataFrame(get_data["id"])
        rows=get_data
        #rows = get_data[columns_quality]
        
        r = 4  # start at fourd row
        c = 1 # column 'a'
        for row in rows:
            #print(row)
            for item in row:
                ws1.cell(row=r, column=c).value = item
                c += 1 # Column 'b'
            c = 1
            r += 1
        #monthly output
        if monthly:
            print ("test monthly ___________________",output)
            list_item_size=output.shape[0]
            ws2=wb["output"]
            
            r = 3  # start at third row
            c = 1 # column 'a'
            for row in range(0,list_item_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
                rows = output.iloc[row]
                for item in rows:
                    ws1.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 1
                r += 1   
            
        
            #filter on non conformity weights
                #part one low weight
            list_item_size=weight_nonconfomity_low.shape[0]
            ws5=wb["wieght_report"]
            rows = weight_nonconfomity_low
            r = 12  # start at 12 row
            c = 1 # column 'a'
            for row in range(0,list_item_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
                rows = weight_nonconfomity_low.iloc[row]
                for item in rows:
                    ws1.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 1
                r += 1   
        
                #part tow hight weight
                ws5=wb["wieght_report"]
            
            
            list_item_size=weight_nonconfomity_high.shape[0]
            rows = weight_nonconfomity_high
            r = 33  # start at 33th row
            c = 1 # column 'a'
            for row in range(0,list_item_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
                rows = weight_nonconfomity_high.iloc[row]
                for item in rows:
                    ws1.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 1
                r += 1   
            
            ws6=wb["ct_report"]
            list_item_size=c_t_nonconfomity.shape[0]
            rows = c_t_nonconfomity
            r = 11  # start at 11th row
            c = 1 # column 'a'
            for row in range(0,list_item_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
                rows = c_t_nonconfomity.iloc[row]
                for item in rows:
                    ws1.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 1
                r += 1   
            
            #filter on on conformity scrap
            
            ws7=wb["scrap_report"]
            list_item_size=scrap_nonconfomity.shape[0]
            rows= scrap_nonconfomity

            r = 15  # start at 15th row
            c = 1 # column 'a'
            for row in range(0,list_item_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
                rows = scrap_nonconfomity.iloc[row]
                for item in rows:
                    ws1.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 1
                r += 1   
            
            #monthly machine report
            ws8=wb["scrap_machine"]
            list_item_size=machines.shape[0]
            rows = machines
            r = 3  # start at 33th row
            c = 1 # column 'a'
            for row in range(0,list_item_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
                rows = machines.iloc[row]
                for item in rows:
                    ws1.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 1
                r += 1   
            '''
            #monthly machine report
            ws8=wb["scrap_machine_yearly"]
            Block.show_machine_yearly_report(self,year,month)
            rows = cursor.fetchall()
            r = 3  # start at 33th row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws8.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            
        
            ws2=wb["output_monthly"]
            Block.show_yearly_report_itemsByMonths(self,year,args)
            
            rows = cursor.fetchall()
            r = 3  # start at third row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws2.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            
            ws3=wb["year"]
            Block.show_machine_report_yearly(self,year,month)
            rows = cursor.fetchall()
            r = 3  # start at third row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws3.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            #scrap monthly
            ws3=wb["month"]
            Block.show_monthly_report_view_month(self,year,month)
            rows = cursor.fetchall()
            r = 3  # start at third row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws3.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            #yearly item report
            ws2=wb["output_yearly"]
            Block.items_report_arabic_custom_item(self,year,month,args)
            
            rows = cursor.fetchall()
            r = 3  # start at third row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws2.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            
            
            #monthly mold report by molds not items
            ws_output_molds=wb["output_molds"]
            Block.show_mnthly_report_molds(self,year,month)

            rows = cursor.fetchall()
            r = 3  # start at third row
            c = 1 # column 'a'
            for row in rows:
                #print(row)

                for item in row:
                    ws_output_molds.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1

            ws_output_molds_yearly=wb["output_mold_monthly"]
            Block.yearly_report_molds_byMonthes(self,year,args)

            rows = cursor.fetchall()
            r = 3  # start at third row
            c = 1 # column 'a'
            for row in rows:
                #print(row)

                for item in row:
                    ws_output_molds_yearly.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1

                #report for moldsy monthly
            
            ws_wieght_yearly=wb["output_molds_yearly"]
            Block.show_yearly_report_molds(self,year,month)
            rows = cursor.fetchall()
            r = 3  # start at 3th row
            c = 1 # column 'a'
            for row in rows:
                for item in row:
                    ws_wieght_yearly.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            
                #report of hight weight
            ws_wieght_yearly=wb["wieght_yearly"]
            Block.yearly_report_ncr_weight(self,year,month)
            rows = cursor.fetchall()
            r = 10  # start at 10th row
            c = 1 # column 'a'
            for row in rows:
                
                for item in row:
                    ws_wieght_yearly.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            #filter on non conformity ct
            ws_ct_yearly=wb["ct_yearly"]
            Block.yearly_report_ncr_ct(self,year,month)
            rows = cursor.fetchall()
            r = 11  # start at 11th row
            c = 1 # column 'a'
            for row in rows:
                
                for item in row:
                    ws_ct_yearly.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            #filter on on conformity scrap
            ws_scrap_yearly=wb["scrap_yearly"]
            Block.yearly_report_ncr_scrap(self,year,month)
            rows= cursor.fetchall()
            r = 15  # start at 15th row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
            
                for item in row:
                    ws_scrap_yearly.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
            #for material
            ws1=wb["materials"]
            Material.materialToPorduct(self,year)
            get_data=cursor.fetchall()
            #get_data.set_index("serial", inplace=True) #put index
            
            #get_data=pd.DataFrame(get_data["id"])
            rows=get_data
            #rows = get_data[columns_quality]
            
            r = 4  # start at fourd row
            c = 1 # column 'a'
            for row in rows:
                #print(row)
                for item in row:
                    ws1.cell(row=r, column=c).value = item
                    c += 1 # Column 'b'
                c = 1
                r += 1
        '''
        wb.save(writerFile)