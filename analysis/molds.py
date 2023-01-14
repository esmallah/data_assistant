'''this module for the data was emported from excel sheet analysis'''

import pandas as pd
import openpyxl as xl
from openpyxl import load_workbook

import numpy as np
import os
from copy import copy
import random
from random import randint,seed
from openpyxl.chart import BarChart, Reference, Series,LineChart
from .collect import columns_quality as qc_molds
#cols=["day","machine","product_name","product_code","product_parts","standard_dry_weight",
#"standard_dry_weight_from","standard_dry_weight_to","average_dry_weight"
#,"standard_rate_hour","c_t_standard_per_second","set","rat_actually","c_t_actually","sum_scrabe_no_parts"
#,"number_scrab_by_item","gross_production_by_set","scrabe_standard"
#,"gross_production","scrap_percent_by_item","customer_id","mold_id","item_id"]
        

        
class Group():
    def __init__(self,folder,readfile,readsheet,column1,column2,writefile,writesheet):
        #self.folder=folder
        self.folder=folder
        self.readfile=readfile
        self.readsheet=readsheet
        self.column1=column1
        self.column2=column2
        self.writefile=writefile
        self.writesheet=writesheet
    def copy_between_workbooks(self):        
        print("________starting copy last sheet____")
        import win32com
        from win32com.client import DispatchEx
        os.chdir(self.folder)
        print("______________test_________")
        print(self.folder)
        #input last day sheets
        
        excel = DispatchEx('Excel.Application')
        wbP=excel.Workbooks.Open('QC_molds_daily_archive')
        wbG=excel.Workbooks.Open(self.readfile)
        #wbP=excel.Workbooks.Open(r'C:\Temp\Junk\Temp.xlsx')
        #wbG=excel.Workbooks.Open(r'C:\Temp\Junk\Temp2.xlsx')
        wbG.Worksheets("deleted").Copy(Before=wbP.Worksheets(self.readsheet))
        wbP.SaveAs(r'C:\Temp\Junk\Temp.xlsx')
        excel.Quit()
        del excel # ensure Excel process ends
    def copy_between_workbooks2(self,worksheet,index=None):        
        print("get last sheet in day")
        os.chdir(self.folder)

        wb=load_workbook(self.readfile)
     #   def add_sheet(self, worksheet, index=None):
        """Add an existing worksheet (at an optional index)."""
#        if not isinstance(worksheet, self._worksheet_class):
 #           raise TypeError("The parameter you have given is not of the type '%s'" % self._worksheet_class.__name__)

        #if index is None:
        #   self.worksheets.append(worksheet)
        #else:
        #self.worksheets.insert(index, worksheet)
        ws_input= wb.get_sheet_by_name(worksheet)
        wb_formats=load_workbook("QC_molds_daily_archive.xlsx")    
        #if not isinstance(worksheet, self._worksheet_class):
        #    raise TypeError("The parameter you have given is not of the type '%s'" % self._worksheet_class.__name__)

        if index is None:
            wb_formats.worksheets.append(ws_input)
        else:
            wb_formats.worksheets.insert(index, worksheet)
        wb_formats.worksheets.insert(1, ws_input)

        wb_formats.save("QC_molds_daily_archive.xlsx")    

    
    def yearly(self):#for statiscs to qc yearly molds report
        os.chdir(self.folder)
        reader=pd.read_excel(self.readfile,self.readsheet)
        #for more filtraation about id_molds
        
        input_report3=reader[qc_molds]
        #input_report2=input_report3[input_report3["customer_name"]==customer]       
        input_report2=input_report3

        input_report=input_report2[input_report2['year']==self.column1]
        #verification 
        #1 for the right ct? ( previous mistaks is bad input)
        #calcuate the logic actioly rat by recalculat form actioly cycltype
        #????????
        #2 for the right names? ( previous mistaks is bad input)
        #calcuate the simulation between names
        #????????
        #ct report
        clos_cycle2=input_report[columns_cycle_time]
        ct_bool=clos_cycle2["c_t_actually"].notnull()
        clos_cycle=clos_cycle2[ct_bool]
        df = pd.DataFrame(clos_cycle, columns = qc_molds )
        rat=pd.crosstab([df.mold_name,df.standard_rate_hour],[df.year,df.month],df.rat_actually,aggfunc='mean',margins=False)
        ct=pd.crosstab([df.mold_name,df.c_t_standard_per_second],[df.year,df.month],df.c_t_actually,aggfunc='mean',margins=False)
        
        #weight report
        clos_weight2=input_report[columns_weight]
        weight_bool=clos_weight2["average_dry_weight"].notnull()
        clos_weight=clos_weight2[weight_bool]
        df2 = pd.DataFrame(clos_weight, columns = qc_molds )

        weight=pd.crosstab([df2.product_name_by_parts,df2.standard_dry_weight,
        df2.standard_dry_weight_from,df2.standard_dry_weight_to],[df2.year,df2.month],df2.average_dry_weight
        ,aggfunc='mean',margins=False)
        
        print("pleas wait")
        #scrap report
        scrap2=input_report[columns_machine]
        scrap_bool=scrap2['gross_production'].notnull()
        scrap=scrap2[scrap_bool]
        
        df3 = pd.DataFrame(scrap, columns = qc_molds )
        
        scrap_set=pd.crosstab([df3.product_name,df3.product_code,df3.scrabe_standard],[df3.year,df3.month],df3.number_scrab_by_item
        ,aggfunc='sum',margins=False)
        
        scrap_parts=pd.crosstab([df3.product_name,df3.product_code,df3.scrabe_standard],[df3.year,df3.month],df3.sum_scrabe_no_parts
        ,aggfunc='sum',margins=False)
        production_set=pd.crosstab([df3.product_name,df3.product_code,df3.scrabe_standard],[df3.year,df3.month],df3.gross_production
        ,aggfunc='sum',margins=False)
        production_parts=pd.crosstab([df3.product_name,df3.product_code,df3.scrabe_standard],[df3.year,df3.month],df3.gross_production
        ,aggfunc='sum',margins=False)
        
        

        #final report
        #day_use=reader["number_day_use"]
        #pd.to_numeric(day_use, errors='ignore')
        #report=clos_cycle.groupby(["mold_name"])["number_day_use"].sum()

        #report["product_name"]=weight["product_name"]
        #report["number_day_use"]=day_use
        
        #ouut put by rename columns
        writer = pd.ExcelWriter(self.writefile)
        input_report=input_report.rename(index=str, columns=col_rename)
        input_report.to_excel(writer,"input",index=False)

        rat=rat.rename(index=str, columns=col_rename)
        rat.to_excel(writer,"rat")
        
        ct=ct.rename(index=str, columns=col_rename)
        ct.to_excel(writer,"ct")
        
        weight=weight.rename(index=str, columns=col_rename)
        weight.to_excel(writer,"weight")
        print("gust wait")
        
        scrap_set=scrap_set.rename(index=str, columns=col_rename)
        scrap_set.to_excel(writer,'scrap_set')
        
        scrap_parts=scrap_parts.rename(index=str, columns=col_rename)
        scrap_parts.to_excel(writer,'scrap_parts')

        production_set=production_set.rename(index=str, columns=col_rename)
        production_set.to_excel(writer,'production_set')
        production_parts=production_parts.rename(index=str, columns=col_rename)
        production_parts.to_excel(writer,'production_parts')
        #report=report.rename(index=str, columns=col_rename)
        #report.to_excel(writer,'report')
        
        clos_weight.to_excel(writer,'input_weight')
        clos_cycle.to_excel(writer,'input_ct')
        scrap.to_excel(writer,'input_scrap')
        writer.save()
    
    def yearly_ct(self):#for statiscs to qc molds report
        os.chdir(self.folder)
        reader=pd.read_excel(self.readfile,self.readsheet)
        #input_report=reader[qc_molds]
        #verification 
        #1 for the right ct? ( previous mistaks is bad input)
        #calcuate the logic actioly rat by recalculat form actioly cycltype
        #????????
        #2 for the right names? ( previous mistaks is bad input)
        #calcuate the simulation between names
        #????????
        #ct report
        clos_cycle2=reader[columns_cycle_time]
        ct_bool=clos_cycle2["c_t_actually"].notnull()
        clos_cycle=clos_cycle2[ct_bool]
        df = pd.DataFrame(clos_cycle, columns = columns_cycle_time )
        rat=pd.crosstab([df.mold_name,df.standard_rate_hour],[df.year,df.month],df.rat_actually,aggfunc='mean',margins=False)
        ct=pd.crosstab([df.mold_name,df.c_t_standard_per_second],[df.year,df.month],df.c_t_actually,aggfunc='mean',margins=False)
        

        #final report
        
        #ouut put by rename columns
        writer = pd.ExcelWriter(self.writefile)
        
        rat=rat.rename(index=str, columns=col_rename)
        rat.to_excel(writer,"rat")
        
        ct=ct.rename(index=str, columns=col_rename)
        ct.to_excel(writer,"ct")
        
        
        
        clos_cycle.to_excel(writer,'input_ct')
        
        writer.save()
    def generate_false_data(self,column1_number,column2_number,row_numbers,cell_form,cell_to,column_numbers):
        os.chdir(self.folder)

        reader=pd.read_excel(self.readfile,self.readsheet)
        
        wb = xl.load_workbook(self.readfile)
        ws= wb.get_sheet_by_name(self.readsheet)
        
        #column_size=reader["month"].shape
        column_size=pd.DataFrame(reader,columns=["month"]).shape[0]
        int(column_size)

        #reader=reader2.iloc[1:column_size+3]#select from 2nd row to last of rows
        #ws2=list(ws.rows)[2]
        
        for i in range(2,column_size):
                       
            LSL=reader.iloc[i][column1_number]#lower limit (row i and column G)
            int(round(LSL,0))
            
            USL=reader.iloc[i][column2_number]#upper limit (row i and column H)
            int(round(USL,0))
            
            weight_average=(LSL+ USL)/2 #average
            differance=(weight_average-LSL)/3 #for decrease diviation
            LCL=weight_average-differance  #for rais CPK
            
            UCL=weight_average+differance  #for rais CPK
              
            
            c = 9 # column 'H'
            for n in range(row_numbers):
                #random.seed(10)
                random_value = random.uniform(LCL,UCL)  
                #random_value = randint(LSL,USL)  
                #str(round(random_value,0))          
                #int(random_value)
                #values.append(random_value)

    #                for item in values:
                
                ws.cell(row=i, column=c).value = random_value
                c += 1
                #n+=1
            
        wb.save(self.writefile)

    def spc_molds(self,year,month,*mold,create_workbook=True):
            '''this functions for select vba anlaysis '''
            print("________starting statsic process charts____")
            #genecal data
            os.chdir(self.folder)
            
            #prepare sheets
            
            
            
            
            #ws_index=wb_formats.copy_worksheet(ws_index)
            save_name=str(year)+"-"+str(month)+" QC_SPC.xlsx"
            #prepare data
            if create_workbook:
                wb_formats=load_workbook("format_QC_reports_v2.xlsx")
                ws_data2=wb_formats.get_sheet_by_name("spc")

                ws_data=wb_formats.copy_worksheet(ws_data2)
                ws_data.title=str(mold)  

            else:
                wb_formats=load_workbook(save_name)
                ws_data2=wb_formats.get_sheet_by_name("spc")

                ws_data=wb_formats.copy_worksheet(ws_data2)
                ws_data.title=str(mold)  
            
            ws_index=wb_formats.get_sheet_by_name("index")
            cols_monthly=["machine_id","mold_name","product_name","product_code","standard_dry_weight",
            "standard_dry_weight_from","standard_dry_weight_to","average_dry_weight","year","month","day",
            "standard_rate_hour","c_t_standard_per_second","c_t_actually","rat_actually","item_id","mold_id","scrabe_standard",'sum_scrabe_shortage_bySet','sum_scrabe_roll_bySet',
            'sum_scrabe_broken_bySet','sum_scrabe_curve_bySet','sum_scrabe_shrinkage_bySet','sum_scrabe_dimentions_bySet','sum_scrabe_weight_bySet','sum_scrabe_dirty_bySet',
            'sum_scrabe_cloration_bySet','sum_scrabe_no_parts','number_scrab_by_item',
            'parts_patchsNumbers','gross_production','scrap_percent_by_item','Items_patchsNumbers','product_by_set_or_no',
            'parts_symbole','part_id','id_DayPartUnique','machine_type','factory','tall_mm','width_mm','deepth_mm','customer_name']

            daily_input= pd.read_excel(self.readfile,"input")
            #Block.show_yearly_report_itemsByMonths(self,year,month)
            #daily_input=cursor.fetchall()
            #data_analysis3=daily_input[cols_monthly]
            last_year=daily_input["year"].max()
            mold_analysis_bool4=daily_input[daily_input["year"]==last_year]
            #mold_analysis3=daily_input[mold_analysis_bool4]
            last_month=mold_analysis_bool4["month"].max()
            
            
            data_analysis3=daily_input[cols_monthly]
            data_analysis2=data_analysis3[data_analysis3["year"]==last_year]
            data_analysis1=data_analysis2[data_analysis2["month"]==last_month]

            year=last_year
            month=last_month
            
            print("_____________year_____",last_year,last_month)            
            #create list of items in monthes
            #list of items
            
            list_item=data_analysis1.groupby(["item_id"])["product_name"].unique()
            
            #list_item = {word: i for i, word in enumerate(items)}
            #list_item=pd.DataFrame()
            #list_item=list_item.append(items)
            list_item_size=pd.DataFrame(list_item,columns=["product_name"]).shape[0]

            #create  the index sheet:
            r = 2  # start at 4th row
            c = 1 # column 'a'
            for row in range(0,list_item_size):  
                rows = list_item.iloc[row]
                for item in rows:
                    ws_index.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 1
                r += 1
            #generate sheets:

            print("items",list_item.index)
            data_analysis=data_analysis1[data_analysis1["item_id"]==mold]

            #for creating data to sheets
            ws_index['g1'] = month            #day for xpscreport
            ws_index['i1'] = year    #month for xps report  
            
            #for i in list_item:
                #ws_copy=wb_formats.copy_worksheet(ws_input)
                #ws_copy.title=str(i)

            #data_analysis=data_analysis1[data_analysis1["item_id"]==254]


            
            #create spc for each mold#________________
            ws_data['g1'] = month            #day for xpscreport
            ws_data['i1'] = year    #month for xps report
            ws_data['c1']=data_analysis["product_name"].iloc[0]
#            ws_data['c1']=str(data_analysis["product_name"].head(1))

            ##analysis weight
            weight2=data_analysis[["day","machine_id","standard_dry_weight",
            "standard_dry_weight_from","standard_dry_weight_to","average_dry_weight"]]
            weight=weight2[weight2["average_dry_weight"].notnull()]
            
            weight_size=pd.DataFrame(weight,columns=["day"]).shape[0]
            r = 18  # start at 4th row
            c = 1 # column 'a'
            for row in range(0,weight_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
                rows = weight.iloc[row]
                for item in rows:
                    ws_data.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 1
                r += 1
            print("_______________test____________")
            print(weight)
            print("weight_size",weight_size)

            ##analysis ct
            ct2=data_analysis[["c_t_standard_per_second"]]
            ct2["c_t_standard_per_second_from"]=ct2["c_t_standard_per_second"]*0.95
            ct2["c_t_standard_per_second_to"]=ct2["c_t_standard_per_second"]*1.05
            ct2["c_t_actually"]=data_analysis["c_t_actually"]
            ct=ct2[ct2["c_t_actually"].notnull()]
            ct_size=pd.DataFrame(ct,columns=["day"]).shape[0]
            
            r = 18  # start at 18th row
            c = 8 # column 'h'
            for row in range(0,ct_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
                rows = ct.iloc[row]
                for item in rows:
                    ws_data.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 8
                r += 1
            ##analysis scrap
            scrap3=data_analysis[["day",'number_scrab_by_item','scrabe_standard','gross_production','scrap_percent_by_item']]
            scrap2=scrap3[scrap3["gross_production"].notnull()]
            scrap=scrap2[scrap2["gross_production"] > scrap2["number_scrab_by_item"]]
            
            #scrap=scrap2
            scrap_size=pd.DataFrame(scrap,columns=["day"]).shape[0]
            
            r = 18  # start at 4th row
            c = 13 # column 'n'
            
            for row in range(0,scrap_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
                rows = scrap.iloc[row]
                for item in rows:
                    ws_data.cell(row=r, column=c).value = item
                    c += 1 # Column 'd'
                c = 13
                r += 1

            #add charts
            #wb_formats.save(str(year)+"-"+str(month)+"MOLD"+str(mold)+" QC_SPC.xlsx")    
            
        ####add chart____________________
            #ws_data=wb_formats.copy_worksheet(ws_spc)
            
            #ws_data=wb_formats.get_sheet_by_name("spc Copy")
            #ws_data.title=str(mold)  
            # create data for plotting 
            chart_weight = LineChart()
            
            #values_weight = Reference(ws_data, min_col = 3, min_row = 18,  max_col = 6, max_row = 40) 
            x_weight = Reference(ws_data, min_col = 1, min_row = 18, max_col = 1, max_row = 40) 
            
            y_weight = Reference(ws_data, min_col = 4, min_row = 18, max_col = 6, max_row = 40) 
            
            #size_weight = Reference(ws_data, min_col = 1, min_row = 18
            # , max_row = 40)
            # create a 1st series of data
            #values_weight = Series(values = y_weight, xvalues = x_weight, zvalues = size_weight , title ="Weight")
            values_weight = Series(values = y_weight, xvalues = x_weight, title ="Weight")
            # Create object of BarChart class 
            
        
            # adding data to the Bar chart object 

            chart_weight.add_data(y_weight) 
            #chart_weight.add_data(x_weight) 
            chart_weight.series.append(values_weight)

            #chart_weight.add_data(values_weight) 
            
            # set the title of the chart 
            chart_weight.title = " dry weight "
            # set the title of the x-axis 
            chart_weight.x_axis.title = " X_AXIS "
            # set the title of the y-axis 
            chart_weight.y_axis.title = " Y_AXIS "
            ws_data.add_chart(chart_weight, "a2") 

            #_____
            #values_ct = Reference(ws_data, min_col = 8, min_row = 17,  max_col = 11, max_row = 40) 
            x_ct = Reference(ws_data, min_col = 9, min_row = 18, max_col = 10, max_row = 40) 
            y_ct = Reference(ws_data, min_col = 11, min_row = 18, max_col = 11, max_row = 40) 
            #size_ct = Reference(ws_data, min_col = 1, min_row = 18, max_row = 40)
            # create a 1st series of data
            values_ct = Series(values = y_ct, xvalues = x_ct, title ="CT")
            # add series data to the chart object
            
            chart_ct = LineChart()
            chart_ct.add_data(x_ct) 
            chart_ct.add_data(y_ct) 
            #chart_ct.series.append(values_ct)

            chart_ct.title = " ct-CHART "
            
            # set the title of the x-axis 
            chart_ct.x_axis.title = " X_AXIS "
            
            # set the title of the y-axis 
            chart_ct.y_axis.title = " Y_AXIS "
            ws_data.add_chart(chart_ct, "h2") 

            #____
            chart_scrap = LineChart()

            #values_scrab = Reference(ws_data, min_col = 13, min_row = 17, max_col = 17, max_row = 40) 
            x_scrab = Reference(ws_data, min_col = 13, min_row = 18, max_col = 11, max_row = 40) 
            y_scrab = Reference(ws_data, min_col = 17, min_row = 18, max_col = 17, max_row = 40) 
            z_scrab = Reference(ws_data, min_col = 15, min_row = 18, max_col = 15, max_row = 40) 
            #size = Reference(ws_data, min_col = 10, min_row = 17, max_row = 40)
            # create a 1st series of data
            #series = Series(values = y_scrab, xvalues = y_scrab, zvalues = size, title ="scrap")
            # add series data to the chart object
            series = Series(values = y_scrab, xvalues = y_scrab, zvalues = z_scrab, title ="scrap")
            #chart_scrap.series.append(series)
            chart_scrap.add_data(x_scrab)
            chart_scrap.add_data(y_scrab)
            chart_scrap.add_data(z_scrab)

            chart_scrap.title = " Scrap-CHART "

            # set the title of the x-axis 
            chart_scrap.x_axis.title = " X_AXIS "
    
            # set the title of the y-axis 
            chart_scrap.y_axis.title = " Y_AXIS "
            ws_data.add_chart(chart_scrap, "o2") 
            # add chart to the sheet 
            # the top-left corner of a chart 
            # is anchored to cell E2 . 
            
                    
                
            #create t
            #for s in wb_formats:
            #values = Reference(scrap)
            #chart = BarChart()
            #chart.add_data(values)
            #ws = wb_formats[ws_spc]
            #ws.add_chart(chart, "B5")
    
        
            wb_formats.save(save_name)