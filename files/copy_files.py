import shutil
import os
import openpyxl as xl
import pandas as pd
from server.config.settings_base import BASE_DIR
class Direcories():
    '''
    for organize files in computer for waste time and updata all related 
    copy image form camira to monthly folder

    copy daily reports to qc share foledr
    '''
    def __init__(self,folder_from,folder_to,file_from,file_to):
        
        self.folder_from=folder_from
        self.folder_to=folder_to
        self.file_from=file_from
        self.file_to=file_to
    def mkdir(self):
        path = (self.folder_from)
        if not os.path.exists(path):
            os.makedirs(path)
    def copy_dir(self):
        print("Before copying file:") 
        path=BASE_DIR
        print(os.listdir(path)) 
        
        # Source path 
        src = self.folder_from
        
        # Destination path 
        dest = self.folder_to
        
        # Copy the content of 
        # source to destination 
        destination = shutil.copytree(src, dest) 
        
        # List files and directories 
        # in "C:/Users / Rajnish / Desktop / GeeksforGeeks" 
        print("After copying file:") 
        print(os.listdir(path)) 
        
        # Print path of newly 
        # created file 
        print("Destination path:", destination)
    def copy_files(self):
        path = (self.folder_from)
        path2= (self.folder_to)
        root_src_dir = os.path.join(path,'source')
        root_target_dir = os.path.join(path2,'target')

        operation = 'copy' # 'copy' or 'move'

        for src_dir, dirs, files in os.walk(root_src_dir):
            dst_dir = src_dir.replace(root_src_dir, root_target_dir)
            if not os.path.exists(dst_dir):
                os.mkdir(dst_dir)
            for file_ in files:
                src_file = os.path.join(src_dir, file_)
                dst_file = os.path2.join(dst_dir, file_)
                if os.path.exists(dst_file):
                    os.remove(dst_file)
                if operation == 'copy':
                    shutil.copy(src_file, dst_dir)
                elif operation == 'move':
                    shutil.move(src_file, dst_dir)
        print(operation,"from ",root_src_dir,"to ",root_target_dir)
    def copy_rename(old_file_name, new_file_name):      #work fine
        src_dir= os.curdir
        dst_dir= os.path.join(os.curdir , "subfolder")
        src_file = os.path.join(src_dir, old_file_name)
        shutil.copy(src_file,dst_dir)
        
        dst_file = os.path.join(dst_dir, old_file_name)
        new_dst_file_name = os.path.join(dst_dir, new_file_name)
        os.rename(dst_file, new_dst_file_name)

        
    def rename(self,old, new):
        new1=["w1","w2","s1","s2"]
        new2=["w1","w1-2","w2","s1","s2"]
        new3=["w1","w1","w2-2","s1","s2"]
        new4=["w1","w1-2","w2-2","s1","s2"]
        dif=[]
        for old in len():
            old=dir[old]
            new=new1[old]
            os.rename(old,new)
    def check_directory(self):
        '''check the name of files in directory as high accuracey for check system files cuase problems'''
        os.chdir(self.folder)

        wb1 = xl.load_workbook(self.readfile1)
        ws1 = wb1.get_sheet_by_name(self.sheet1) 

        wb2 = xl.load_workbook(self.writefile)
        ws2 = wb2.get_sheet_by_name(self.writesheet) 
        for root,dirs,files in os.walk('.',topdown=False):        
            for name in files:
                print(os.path.join(root, name))
            for name in dirs:
                print(os.path.join(root, name))
class Select():
    """this class provide  work books and sheet names as input """
    def __init__(self,folder,readfile1,sheet1,writefile,writesheet):
        self.folder=folder
        
        self.readfile1=readfile1
        self.sheet1=sheet1
        
        self.writefile=writefile
        self.writesheet=writesheet

    def copy_workbooks(self):
        '''for copy ranges form multi workbooks to another sheer'''
        os.chdir(self.folder)
        all_data = pd.DataFrame()
        for root,dirs,files in os.walk('.',topdown=False):        
            for name in files:
                print(os.path.join(root, name))
            for name in dirs:      
                df = pd.read_excel(name,self.sheet1)
                df["month"]=name
                all_data = all_data.append(df,ignore_index=True)

        # now save the data frame
        writer = pd.ExcelWriter(self.writefile)
        all_data.to_excel(writer,self.writesheet)
        writer.save()    
                                    
    def copy_sheet(self):
        '''for copy range form 1sheet to another sheer'''
        os.chdir(self.folder)
        
        
        print(os.chdir)
    
        wb1 = xl.load_workbook(self.readfile1)
        ws1 = wb1.get_sheet_by_name(self.sheet1) 

        wb2 = xl.load_workbook(self.writefile)
        ws2 = wb2.get_sheet_by_name(self.writesheet) 

        for row in ws1:
            for cell in row:
                ws2[cell.coordinate].value = cell.value

        wb2.save(self.writefile)
    def multi_sheet(self):
        '''for append all sheets in the same workbook by indix handling by columns name'''
        os.chdir(self.folder)
        print(os.chdir)
        
        all_data = pd.DataFrame()
        days=[1,2,3,4,5,6,7,8,9,10,11,12]
        for f in days:
            df = pd.read_excel(self.readfile1,f)
            all_data = all_data.append(df,ignore_index=True)

        # now save the data frame
        writer = pd.ExcelWriter(self.writefile)
        all_data.to_excel(writer,self.writesheet)
        writer.save()    
        
    def multi_workbook(self):
        '''for append all workbooks in the same directory  by indix handling by columns name
        '''
        os.chdir(self.folder)
        all_data = pd.DataFrame()
        for f in glob.glob("*."+self.readfile1):
            df = pd.read_excel(f,self.sheet1)
            df["month2"]=f[0:2]
            all_data = all_data.append(df,ignore_index=False)
        # now save the data frame
        writer = pd.ExcelWriter(self.writefile)
        all_data.to_excel(writer,self.writesheet)
        writer.save()    
    def re_right_sheet(self):
        os.chdir(self.folder)
        wb = xl.load_workbook(self.readfile1)
        #year=
        #month=

        #daily input
        rows=wb[self.sheet1]
        r = 3  # start at third row
        c = 1 # column 'a'
        for row in rows:
            #print(row)

            for item in row:
                ws2.cell(row=r, column=c).value = item
                c += 1 # Column 'b'
            c = 1
            r += 1
        
        wb.save("electolox_list.xlsx")
    def re_right_sheet2(self):
        os.chdir(self.folder)
        wb = pd.read_excel(self.readfile1,self.sheet1)
        writer = pd.ExcelWriter(self.writefile)
        wb.to_excel(writer,self.writesheet, index=False)