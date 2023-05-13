import os
import openpyxl
from openpyxl import Workbook
import pandas as pd


wb = openpyxl.Workbook()
ws = wb.active

class Files_control():
    def __init__(self,folder,folder2):
        self.folder=folder
        self.folder2=folder2
    def get_Files_names (path,outputpath): 
        list_files=[]
        list_path=[]
        list_filepath=[]

        last_folder=[]
        os.chdir(outputpath)
        current_file = os.path.abspath(os.path.dirname(__file__))
        import pathlib
            
        with open("F01-INS-QES-P-02_documentation_master_list.xls", "w") as a:
            for filepath, subdirs, files in os.walk(path):
                for filename in files:
                    f = os.path.join( filename)
          #          d = os.path.join( subdirs)
                    list_files.append(str(f))
#                    if x.endswith(".shp"):
                    list_path.append(os.path.join(filepath))
                    list_filepath.append(os.path.join(filepath, filename))
        #            get_data.append(str(f),os.path.join(path, filename))
#                    p = pathlib.Path(filepath)
                    #remove first folder
                    #p.parts[2:]
                    #path2=pathlib.Path(*p.parts[2:])
                    
                    #select last folder
                    number_folders=len(os.path.split(filepath))
#                    max_folder=max(number_folders)
                    
                    last_folder.append(os.path.split(filepath[0]))

            get_data=pd.DataFrame({"category":last_folder,"files":list_files,"path":list_path,"link":list_filepath},)
            get_data['filename'] = get_data['files'].str.rstrip('.')
            print("___________test_____________",number_folders)
            
                        
                    #get_data.append(os.path.join(path, filename))        
        print(list_files[10])

    #    f= open('list.xlsx','w')  
    #   for index,filename in enumerate(list_files):
    #      f.write("%s. %s \n"%(index,filename.encode("utf-8")))
        
        ws.title = "files"
        list_item_size=get_data.shape[0]
        ws1=wb["files"]
        #create  the index sheet:
        r = 4  # start at 4th row
        c = 1 # column 'a'
        for row in range(0,list_item_size):       #you must start by 0 to catch all data , if you start by 1 you ignore first row in data source
            rows = get_data.iloc[row]
            for item in rows:
                ws1.cell(row=r, column=c).value = item
        #        ws1.cell(row=r, column=c).style = "Hyperlink"
                ws1.cell(row=r, column=c).hyperlink = item
                c += 1 # Column 'd'
            c = 1
            r += 1   
        '''
        r = 2  # start at fourd row
        c = 1 # column 'a'
        #for item in list_path:
        for item in df_location:
            ws.cell(row=r, column=c).value = item
            r += 1 # Column 'b'
        r = 2
        c += 1
        '''
        wb.save("list.xlsx")
        return list_files

    #2 write xcel file for folders name
    
        
    def get_folders_list (path,outputpath): 
        os.chdir(outputpath)
        folders= os.listdir (path) # get all files' and folders' names in the current directory
        result = ['the departments']
        for folder in folders: # loop through all the files and folders
            if os.path.isdir(os.path.join(os.path.abspath(path), folder)): # check whether the current object is a folder or not
                result.append(folder) 
        result.sort()
        f= open('list.xlsx','w')
        for index,folder in enumerate(result):
            f.write("%s. %s \n"%(index,folder))
        
        f.close()

    def resize_image(path):
        #!/usr/bin/python
        from PIL import Image
        import os, sys

        #path = "/root/Desktop/python/images/"
        dirs = os.listdir( path )

        def resize():
            for item in dirs:
                if os.path.isfile(path+item):
                    im = Image.open(path+item)
                    f, e = os.path.splitext(path+item)
                    imResize = im.resize((200,200), Image.ANTIALIAS)
                    imResize.save(f + ' resized.jpg', 'JPEG', quality=90)

        resize()
    def creatfolders(self,path,outputpath):
        os.chdir(path)
        
        if not os.path.exists(outputpath):
            os.makedirs(outputpath)